import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import numpy as np
import re
import unicodedata

def parse_ntd_file(uploaded_file):
    """
    Giải mã file .NTD: đọc POLE (tim tuyến) và TARGETL/TARGETR (cánh)
    """
    data_points = []
    raw_content = uploaded_file.read()
    try:
        lines = raw_content.decode("utf-8").splitlines()
    except UnicodeDecodeError:
        lines = raw_content.decode("latin1").splitlines()
        
    current_x = 0.0
    current_pole = ""
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        parts = line.split()
        if not parts:
            continue
            
        token = parts[0].upper()
        
        if token == 'POLE' and len(parts) >= 4:
            try:
                current_pole = parts[1].strip().upper()
                current_x = float(parts[2])
                z_tim = float(parts[3])
                data_points.append({
                    'Cọc': current_pole, 'Lý trình': current_x, 'Offset': 0.0, 'Z': z_tim, 'Tag_Gốc': 'POLE'
                })
            except ValueError:
                pass
        elif token in ['TARGETL', 'TARGETR'] and len(parts) >= 3:
            try:
                dist_offset = float(parts[1])
                z_val = float(parts[2])
                if current_pole:
                    data_points.append({
                        'Cọc': current_pole, 'Lý trình': current_x, 'Offset': dist_offset, 'Z': z_val, 'Tag_Gốc': token
                    })
            except ValueError:
                pass
                
    return pd.DataFrame(data_points)

def parse_coordinate_file(uploaded_file):
    """
    Đọc bảng tọa độ VN-2000 (Excel hoặc CSV)
    """
    try:
        if uploaded_file.name.endswith('.csv'):
            df_coord = pd.read_csv(uploaded_file, skiprows=1)
        else:
            df_coord = pd.read_excel(uploaded_file, skiprows=1)
            
        df_coord.columns = [str(c).strip().upper() for c in df_coord.columns]
        
        try:
            col_name = [c for c in df_coord.columns if 'CỌC' in c or 'TEN' in c][0]
        except IndexError:
            col_name = df_coord.columns[1]
            
        col_x = df_coord.columns[3]
        col_y = df_coord.columns[4]
        
        x_numeric = pd.to_numeric(df_coord[col_x], errors='coerce')
        y_numeric = pd.to_numeric(df_coord[col_y], errors='coerce')
        
        df_clean = pd.DataFrame({
            'Cọc_Excel': df_coord[col_name].astype(str).str.strip().str.upper(),
            'X_VN2000': x_numeric,
            'Y_VN2000': y_numeric
        })
        return df_clean.dropna(subset=['X_VN2000', 'Y_VN2000']).reset_index(drop=True)
    except Exception as e:
        st.error(f"Lỗi đọc file bảng tọa độ VN-2000: {e}")
        return None

def convert_to_vn2000(df_ntd, df_coord):
    """
    Đồng bộ tọa độ thực tế cho các điểm đo theo tim tuyến
    """
    try:
        list_ntd_x = sorted(df_ntd['Lý trình'].unique())
        df_coord_clean = df_coord.copy()
        
        min_len = min(len(list_ntd_x), len(df_coord_clean))
        if min_len == 0:
            return pd.DataFrame()
            
        map_x_real = {}
        map_y_real = {}
        for i in range(min_len):
            ly_trinh_ntd = list_ntd_x[i]
            map_x_real[ly_trinh_ntd] = df_coord_clean['X_VN2000'].iloc[i]
            map_y_real[ly_trinh_ntd] = df_coord_clean['Y_VN2000'].iloc[i]
            
        df_merged = df_ntd.copy()
        df_merged['X_VN2000'] = df_merged['Lý trình'].map(map_x_real)
        df_merged['Y_VN2000'] = df_merged['Lý trình'].map(map_y_real)
        
        df_merged = df_merged.dropna(subset=['X_VN2000', 'Y_VN2000']).copy()
        
        df_tim_calc = df_merged[df_merged['Offset'] == 0].drop_duplicates(subset=['Lý trình']).sort_values('Lý trình').copy()
        
        if len(df_tim_calc) >= 2:
            df_tim_calc['dX'] = np.gradient(df_tim_calc['X_VN2000'].values)
            df_tim_calc['dY'] = np.gradient(df_tim_calc['Y_VN2000'].values)
        else:
            df_tim_calc['dX'] = 1.0
            df_tim_calc['dY'] = 0.0
            
        df_tim_calc['Góc_Tuyến'] = np.arctan2(df_tim_calc['dY'], df_tim_calc['dX'])
        
        goc_map = dict(zip(df_tim_calc['Lý trình'], df_tim_calc['Góc_Tuyến']))
        df_merged['Góc_Tuyến'] = df_merged['Lý trình'].map(goc_map).bfill().ffill()
        
        angle_offset = df_merged['Góc_Tuyến'] + (np.pi / 2)
        df_merged['X_Real'] = df_merged['X_VN2000'] + df_merged['Offset'] * np.cos(angle_offset)
        df_merged['Y_Real'] = df_merged['Y_VN2000'] + df_merged['Offset'] * np.sin(angle_offset)
        
        return df_merged
    except Exception as e:
        st.error(f"Lỗi xử lý đồng bộ chuỗi điểm tim thực địa: {e}")
        return pd.DataFrame()

def ve_dia_hinh_3d(df, he_so_z=1.0, che_do="Bề mặt mịn", do_min=3,
                   x_origin: float = 0.0, y_origin: float = 0.0):
    if df.empty:
        return None, None, None, None
    try:
        df_clean = df.sort_values(['Lý trình', 'Offset']).copy()
        unique_lts = sorted(df_clean['Lý trình'].unique())
        
        # Mở rộng 50m mỗi đầu
        lt_min = unique_lts[0]
        lt_max = unique_lts[-1]
        lt_left = lt_min - 50.0
        lt_right = lt_max + 50.0

        df_first = df_clean[df_clean['Lý trình'] == lt_min].sort_values('Offset')
        df_last = df_clean[df_clean['Lý trình'] == lt_max].sort_values('Offset')

        def create_extended_section(df_template, new_lt, delta_lt):
            goc = df_template['Góc_Tuyến'].iloc[0]
            ux = np.cos(goc)
            uy = np.sin(goc)
            new_df = df_template.copy()
            new_df['Lý trình'] = new_lt
            new_df['X_Real'] = new_df['X_Real'] + delta_lt * ux
            new_df['Y_Real'] = new_df['Y_Real'] + delta_lt * uy
            new_df['Cọc'] = new_df['Cọc'] + f"_ext{new_lt:.0f}"
            return new_df

        df_left = create_extended_section(df_first, lt_left, lt_left - lt_min)
        df_right = create_extended_section(df_last, lt_right, lt_right - lt_max)

        df_clean = pd.concat([df_left, df_clean, df_right], ignore_index=True)
        # Normalize VN-2000 absolute coords → relative (lý trình system)
        if x_origin != 0.0 and 'X_Real' in df_clean.columns:
            df_clean['X_Real'] = df_clean['X_Real'] - x_origin
        if y_origin != 0.0 and 'Y_Real' in df_clean.columns:
            df_clean['Y_Real'] = df_clean['Y_Real'] - y_origin
        df_clean = df_clean.sort_values(['Lý trình', 'Offset']).reset_index(drop=True)
        unique_lts = sorted(df_clean['Lý trình'].unique())

        num_samples = 40
        target_pct = np.linspace(0.0, 1.0, num_samples)
        matrix_x, matrix_y, matrix_z = [], [], []
        
        # Xác định cọc có TARGET hợp lệ (≥ 2 điểm)
        target_lts = []
        for lt in unique_lts:
            df_sub = df_clean[df_clean['Lý trình'] == lt]
            if df_sub['Tag_Gốc'].str.contains('TARGET', na=False).sum() >= 2:
                target_lts.append(lt)
        
        if not target_lts:
            st.error("Không có cọc TARGET hợp lệ")
            return None, None, None, None
        
        for lt in unique_lts:
            df_sub = df_clean[df_clean['Lý trình'] == lt].sort_values('Offset')
            df_target = df_sub[df_sub['Tag_Gốc'].str.contains('TARGET', na=False)]
            if len(df_target) >= 2:
                obs_offsets = df_target['Offset'].values
                obs_x_real = df_target['X_Real'].values
                obs_y_real = df_target['Y_Real'].values
                obs_zs = df_target['Z'].values
            else:
                # Tìm cọc TARGET gần nhất
                near_lt = None
                left = [t for t in target_lts if t < lt]
                if left:
                    near_lt = max(left)
                else:
                    right = [t for t in target_lts if t > lt]
                    if right:
                        near_lt = min(right)
                if near_lt is None:
                    continue
                df_near = df_clean[df_clean['Lý trình'] == near_lt].sort_values('Offset')
                df_near_target = df_near[df_near['Tag_Gốc'].str.contains('TARGET', na=False)]
                if len(df_near_target) < 2:
                    continue
                delta = lt - near_lt
                goc = df_near['Góc_Tuyến'].iloc[0]
                ux = np.cos(goc)
                uy = np.sin(goc)
                obs_offsets = df_near_target['Offset'].values
                obs_x_real = df_near_target['X_Real'].values + delta * ux
                obs_y_real = df_near_target['Y_Real'].values + delta * uy
                obs_zs = df_near_target['Z'].values
            
            # Sắp xếp theo offset
            idx = np.argsort(obs_offsets)
            obs_offsets = obs_offsets[idx]
            obs_x_real = obs_x_real[idx]
            obs_y_real = obs_y_real[idx]
            obs_zs = obs_zs[idx]
            if obs_offsets[-1] - obs_offsets[0] < 1e-6:
                x_line = np.full(num_samples, obs_x_real[0])
                y_line = np.full(num_samples, obs_y_real[0])
                z_line = np.full(num_samples, obs_zs[0])
            else:
                pct = (obs_offsets - obs_offsets[0]) / (obs_offsets[-1] - obs_offsets[0])
                x_line = np.interp(target_pct, pct, obs_x_real)
                y_line = np.interp(target_pct, pct, obs_y_real)
                z_line = np.interp(target_pct, pct, obs_zs)
            matrix_x.append(x_line)
            matrix_y.append(y_line)
            matrix_z.append(z_line)
        
        matrix_x = np.array(matrix_x)
        matrix_y = np.array(matrix_y)
        matrix_z = np.array(matrix_z)
        
        # Nội suy toàn bộ NaN dọc theo từng cột (theo chiều dọc lý trình)
        df_z = pd.DataFrame(matrix_z)
        df_z = df_z.interpolate(method='linear', axis=0, limit_direction='both')
        matrix_z = df_z.values
        
        # Làm mịn (rolling mean)
        if do_min > 1:
            mz_pd = pd.DataFrame(matrix_z)
            mz_pd = mz_pd.rolling(window=do_min, min_periods=1, center=True).mean()
            matrix_z = mz_pd.T.rolling(window=do_min, min_periods=1, center=True).mean().T.values
            # Nếu phát sinh NaN, nội suy lại
            if np.isnan(matrix_z).any():
                df_z2 = pd.DataFrame(matrix_z)
                df_z2 = df_z2.interpolate(method='linear', axis=0, limit_direction='both')
                matrix_z = df_z2.values
        
        z_scaled = matrix_z * he_so_z
        fig = go.Figure()
        
        # Dùng hovertemplate với %{z:.2f} thay vì customdata
        if che_do in ["Bề mặt mịn", "Lưới tam giác"]:
            show_wireframe = (che_do == "Lưới tam giác")
            # name + showlegend + legendgroup → ẨN/HIỆN ĐỊA HÌNH bằng chú giải
            # (trước đây mặt địa hình không khai báo gì nên không tắt được, che
            #  mất kết cấu bên dưới khi muốn xem riêng cầu).
            fig.add_trace(go.Surface(
                x=matrix_x, y=matrix_y, z=z_scaled,
                colorscale='Earth', reversescale=True, opacity=0.95,
                colorbar=dict(title=dict(text="Cao độ Z (m)", side="right"), thickness=15),
                name="Địa hình", legendgroup="Địa hình", showlegend=True,
                hovertemplate="X: %{x:.1f} m<br>Y: %{y:.1f} m<br>Z: %{z:.2f} m<extra></extra>",
                contours=dict(
                    x=dict(show=show_wireframe, color="rgba(0,0,0,0.2)", width=1),
                    y=dict(show=show_wireframe, color="rgba(0,0,0,0.2)", width=1)
                )
            ))
        elif che_do == "Đường đồng mức":
            fig.add_trace(go.Surface(
                x=matrix_x, y=matrix_y, z=z_scaled,
                colorscale='Viridis', opacity=0.95,
                colorbar=dict(title=dict(text="Cao độ Z (m)", side="right"), thickness=15),
                name="Địa hình", legendgroup="Địa hình", showlegend=True,
                contours_z=dict(show=True, usecolormap=False, color="rgb(0,0,0)", width=2, project=dict(z=True)),
                hovertemplate="X: %{x:.1f}<br>Y: %{y:.1f}<br>Z: %{z:.2f} m<extra></extra>"
            ))
        
        # Đường tim tuyến
        df_tim_all = df_clean[df_clean['Offset'] == 0].drop_duplicates(subset=['Lý trình']).sort_values('Lý trình')
        if not df_tim_all.empty:
            nhan_hien_thi = df_tim_all.apply(lambda r: f"{r['Cọc']} (LT: {r['Lý trình']:.1f}m)", axis=1).values
            fig.add_trace(go.Scatter3d(
                x=df_tim_all['X_Real'].values, y=df_tim_all['Y_Real'].values, z=df_tim_all['Z'].values * he_so_z,
                mode='lines+markers+text',
                line=dict(color='red', width=4), marker=dict(size=4, color='yellow'),
                text=nhan_hien_thi, textposition="top center",
                textfont=dict(family="Arial, sans-serif", size=11, color="lightblue"),
                name='Đường tim tuyến dọc sông'
            ))
        
        fig.update_layout(
            title=dict(text="🏔️ MÔ HÌNH ĐỊA HÌNH KHÔNG GIAN 3D BÁM SÁT TOÀN TUYẾN NTD", font=dict(size=16, color='#007acc')),
            height=850,
            scene=dict(
                xaxis_title="Tọa độ X VN-2000 (m)",
                yaxis_title="Tọa độ Y VN-2000 (m)",
                zaxis_title="Cao độ Z (m)",
                aspectmode='data'
            ),
            template="plotly_dark",
            margin=dict(l=10, r=10, t=40, b=10),
            paper_bgcolor='rgba(0,0,0,0)'
        )
        return fig, matrix_x, matrix_y, matrix_z
    except Exception as e:
        st.error(f"Lỗi phân tích đồ họa không gian: {e}")
        return None, None, None, None

# ============================================================================
# XỬ LÝ ĐỊA CHẤT
# ============================================================================

def doc_excel_dia_chat_3_sheet(uploaded_file):
    """
    Đọc file Excel địa chất (sheet tọa độ + các sheet lớp đất)
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(uploaded_file, data_only=True)
        sheet_names = wb.sheetnames

        # Tìm sheet tọa độ
        sheet_toado = None
        for s in sheet_names:
            if 'toado' in s.lower().replace('_', '').replace(' ', ''):
                sheet_toado = s
                break
        if not sheet_toado:
            st.error("Không tìm thấy sheet tọa độ (chứa 'Toado')")
            return None, None, None

        ws = wb[sheet_toado]
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        df_hk_raw = pd.DataFrame(data, columns=headers)
        df_hk_raw.columns = [str(c).strip().upper().replace(' ', '_') for c in df_hk_raw.columns]

        def find_col(cols, candidates):
            for col in cols:
                for cand in candidates:
                    if cand in col:
                        return col
            return None

        col_name = find_col(df_hk_raw.columns, ['HO_KHOAN', 'HỐ_KHOAN'])
        col_x = find_col(df_hk_raw.columns, ['X', 'X_VN2000'])
        col_y = find_col(df_hk_raw.columns, ['Y', 'Y_VN2000'])
        col_z = find_col(df_hk_raw.columns, ['Z', 'Z_MIENG'])

        if not (col_name and col_x and col_y):
            st.error(f"Thiếu cột tên hố, X hoặc Y. Các cột: {list(df_hk_raw.columns)}")
            return None, None, None

        df_hk = pd.DataFrame({
            'Ho_Khoan': df_hk_raw[col_name].astype(str).str.strip().str.upper(),
            'X_VN2000': pd.to_numeric(df_hk_raw[col_x], errors='coerce'),
            'Y_VN2000': pd.to_numeric(df_hk_raw[col_y], errors='coerce'),
            'Z_Mieng': pd.to_numeric(df_hk_raw[col_z], errors='coerce') if col_z else 0.0
        }).dropna(subset=['X_VN2000', 'Y_VN2000']).reset_index(drop=True)

        # Đọc các sheet lớp đất
        list_layers = []
        for sheet_name in sheet_names:
            if sheet_name.lower() == sheet_toado.lower() or 'spt' in sheet_name.lower():
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if not headers:
                continue
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    data_rows.append(row)
            if not data_rows:
                continue
            df = pd.DataFrame(data_rows, columns=headers)
            df.columns = [str(c).strip().upper().replace(' ', '_') for c in df.columns]

            col_ten = find_col(df.columns, ['TEN_LOP', 'TÊN_LỚP'])
            col_day = find_col(df.columns, ['CAO_DO_DAY_LOP', 'CAO_DO_DAY', 'DAY_LOP'])
            if not (col_ten and col_day):
                continue

            ten_hk = sheet_name.strip().upper()
            for _, row in df.iterrows():
                ten_lop = str(row[col_ten]).strip().upper()
                if not ten_lop or ten_lop == 'NAN':
                    continue
                try:
                    cao_do_day = float(row[col_day]) if row[col_day] is not None else np.nan
                except:
                    cao_do_day = np.nan
                if np.isnan(cao_do_day):
                    continue
                list_layers.append({
                    'Ho_Khoan': ten_hk,
                    'Ten_Lop': ten_lop,
                    'Cao_Do_Day': cao_do_day
                })

        df_layers = pd.DataFrame(list_layers)

        # Đọc SPT nếu có
        df_spt = None
        sheet_spt = [s for s in sheet_names if 'spt' in s.lower()]
        if sheet_spt:
            df_spt = pd.read_excel(uploaded_file, sheet_name=sheet_spt[0], header=0)
            df_spt.columns = [str(c).strip().upper().replace(' ', '_') for c in df_spt.columns]

        return df_hk, df_layers, df_spt

    except Exception as e:
        st.error(f"Lỗi đọc file Excel: {e}")
        return None, None, None

def _lam_sach_ten_hk(text):
    s = str(text).strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Za-z0-9]', '', s)
    return s.upper()

def _tinh_tuyen_dia_hinh(matrix_x, matrix_y):
    mid = matrix_x.shape[1] // 2
    cx = matrix_x[:, mid]
    cy = matrix_y[:, mid]
    dx = np.diff(cx)
    dy = np.diff(cy)
    chainage = np.zeros(len(cx))
    chainage[1:] = np.cumsum(np.sqrt(dx**2 + dy**2))
    x0, y0 = cx[0], cy[0]
    x1, y1 = cx[-1], cy[-1]
    vlen = np.sqrt((x1 - x0)**2 + (y1 - y0)**2) + 1e-6
    return chainage, (x1 - x0)/vlen, (y1 - y0)/vlen, x0, y0

def _chainage_diem(x, y, ux, uy, x0, y0):
    return (x - x0)*ux + (y - y0)*uy

def _xay_dung_ho_so_lop(df_hk_v, df_layers, lop_dat, ext_m=50.0):
    """
    Xây dựng profile đáy lớp dọc tuyến.
    - Nếu có >=2 hố: nội suy tuyến tính, kéo dài ext_m (mặc định 50m).
    - Nếu chỉ có 1 hố:
        * Tìm hố lân cận gần nhất bên trái và phải (nếu có)
        * Kéo dài đến trung điểm với hố lân cận đó
        * Nếu không có lân cận (đầu hoặc cuối tuyến) thì kéo dài 25m
    """
    pt_c, pt_z = [], []
    for _, hk in df_hk_v.iterrows():
        sub = df_layers[(df_layers['Ho_Khoan'] == hk['Ho_Khoan']) & (df_layers['Ten_Lop'] == lop_dat)]
        if sub.empty:
            continue
        cao_day = sub['Cao_Do_Day'].iloc[0]
        pt_c.append(hk['Chainage'])
        pt_z.append(cao_day)
    
    if len(pt_c) == 0:
        return None, None
    
    idx = np.argsort(pt_c)
    c_arr = np.array(pt_c)[idx]
    z_arr = np.array(pt_z)[idx]
    
    # Trường hợp chỉ có 1 hố
    if len(c_arr) == 1:
        c0 = c_arr[0]
        z0 = z_arr[0]
        all_c = df_hk_v['Chainage'].values
        left = None
        right = None
        for c in all_c:
            if c < c0:
                if left is None or c > left:
                    left = c
            elif c > c0:
                if right is None or c < right:
                    right = c
        left_bound = (c0 + left) / 2 if left is not None else c0 - 25.0
        right_bound = (c0 + right) / 2 if right is not None else c0 + 25.0
        c_arr = np.array([left_bound, c0, right_bound])
        z_arr = np.array([z0, z0, z0])
    else:
        # Có >=2 hố
        c_arr = np.insert(c_arr, 0, c_arr[0] - ext_m)
        z_arr = np.insert(z_arr, 0, z_arr[0])
        c_arr = np.append(c_arr, c_arr[-1] + ext_m)
        z_arr = np.append(z_arr, z_arr[-1])
    
    return c_arr, z_arr

def _mask_bien_xy_dia_hinh(matrix_x, matrix_y):
    nr, nc = matrix_x.shape
    bien_trai = np.column_stack([matrix_x[:,0], matrix_y[:,0]])
    bien_phai = np.column_stack([matrix_x[:,-1], matrix_y[:,-1]])[::-1]
    bien_duoi = np.column_stack([matrix_x[-1,:], matrix_y[-1,:]])
    bien_tren = np.column_stack([matrix_x[0,:], matrix_y[0,:]])[::-1]
    polygon = np.vstack([bien_trai, bien_duoi, bien_phai, bien_tren])
    xq = matrix_x.ravel()
    yq = matrix_y.ravel()
    inside = np.zeros(xq.shape, dtype=bool)
    n = len(polygon)
    for i in range(n):
        x1, y1 = polygon[i]
        x2, y2 = polygon[(i+1)%n]
        cond = ((y1 > yq) != (y2 > yq)) & (xq < (x2 - x1)*(yq - y1)/((y2 - y1)+1e-9) + x1)
        inside ^= cond
    return inside.reshape(matrix_x.shape)

def _dag_luoi_mat_lop(chainage_rows, c_arr, z_arr, shape):
    grid_z = np.full(shape, np.nan)
    for i, c_row in enumerate(chainage_rows):
        grid_z[i,:] = np.interp(c_row, c_arr, z_arr)
    return grid_z

def _cat_lop_duoi_dia_hinh(grid_z, matrix_z, eps=0.05):
    out = grid_z.copy()
    valid = ~np.isnan(out) & ~np.isnan(matrix_z)
    mask_cat = valid & (out > matrix_z)
    out[mask_cat] = matrix_z[mask_cat] - eps
    return out

def _thu_tu_lop_dat(df_layers):
    col_ten = None
    for col in df_layers.columns:
        if 'TEN_LOP' in col.upper():
            col_ten = col
            break
    if col_ten is None:
        return []
    col_day = None
    for col in df_layers.columns:
        if 'CAO_DO_DAY' in col.upper():
            col_day = col
            break
    if col_day is None:
        return []
    depth_mean = df_layers.groupby(col_ten)[col_day].mean()
    return depth_mean.sort_values(ascending=False).index.tolist()

def dap_them_ket_cau_dia_chat_3d(fig, df_hk, df_layers, df_spt, matrix_x, matrix_y, matrix_z, he_so_z=1.0,
                                 hien_mat_phang_lop=True, hien_khoi_lop=False, do_trong_dia_hinh=0.72):
    if fig is None or df_hk is None or df_hk.empty or df_layers is None or df_layers.empty:
        return fig
    if matrix_x is None or matrix_y is None or matrix_z is None:
        return fig

    df_hk_clean = df_hk.copy()
    df_layers_clean = df_layers.copy()
    df_hk_clean['Key_HK'] = df_hk_clean['Ho_Khoan'].apply(_lam_sach_ten_hk)
    df_layers_clean['Key_HK'] = df_layers_clean['Ho_Khoan'].apply(_lam_sach_ten_hk)

    chainage_rows, ux, uy, x0, y0 = _tinh_tuyen_dia_hinh(matrix_x, matrix_y)
    df_hk_clean['Chainage'] = df_hk_clean.apply(
        lambda r: _chainage_diem(r['X_VN2000'], r['Y_VN2000'], ux, uy, x0, y0), axis=1
    )
    df_hk_v = df_hk_clean.sort_values('Chainage').reset_index(drop=True)

    danh_sach_lop = _thu_tu_lop_dat(df_layers_clean)
    mau_sac = ['#d73027', '#f46d43', '#fdae61', '#fee090', '#e0f3f8', '#abd9e9', '#74add1', '#4575b4', '#313695', '#a6d96a']

    for idx, lop in enumerate(danh_sach_lop):
        # Lấy danh sách hố khoan có lớp này
        hks = df_layers_clean[df_layers_clean['Ten_Lop'] == lop]['Ho_Khoan'].unique()
        so_hk = len(hks)
        
        # Tạo profile (c_arr, z_arr) – vẫn dùng hàm cũ (có kéo dài)
        c_arr, z_arr = _xay_dung_ho_so_lop(df_hk_v, df_layers_clean, lop, ext_m=50.0)
        if c_arr is None:
            continue
        
        # Nội suy ra lưới
        grid_z = _dag_luoi_mat_lop(chainage_rows, c_arr, z_arr, matrix_x.shape)
        mask_xy = _mask_bien_xy_dia_hinh(matrix_x, matrix_y)
        grid_z[~mask_xy] = np.nan
        
        # --- XỬ LÝ THEO SỐ LƯỢNG HỐ ---
        if so_hk == 1:
            # Lấy chainage của hố duy nhất
            hk_chainage = df_hk_v[df_hk_v['Ho_Khoan'].isin(hks)]['Chainage'].values[0]
            radius = 15.0  # Bán kính ảnh hưởng (m), có thể sửa thành 10, 20 tùy ý
            mask_cuc_bo = (chainage_rows >= hk_chainage - radius) & (chainage_rows <= hk_chainage + radius)
            grid_z[~mask_cuc_bo, :] = np.nan
        else:
            # Với >=2 hố, giới hạn trong khoảng profile (tránh tràn ra ngoài)
            min_c = np.min(c_arr)
            max_c = np.max(c_arr)
            mask_profile = (chainage_rows >= min_c) & (chainage_rows <= max_c)
            grid_z[~mask_profile, :] = np.nan
        
        grid_z_scaled = grid_z * he_so_z
        if np.all(np.isnan(grid_z)):
            continue
        
        if hien_mat_phang_lop:
            fig.add_trace(go.Surface(
                x=matrix_x, y=matrix_y, z=grid_z_scaled,
                colorscale=[[0, mau_sac[idx % len(mau_sac)]], [1, mau_sac[idx % len(mau_sac)]]],
                opacity=0.65, showscale=False,
                name=f"Đáy lớp {lop}",
                hovertemplate=f"Lớp {lop}<br>X: %{{x:.1f}}<br>Y: %{{y:.1f}}<br>Z đáy: %{{z:.2f}}<extra></extra>"
            ))
            
    return fig
import uuid as _uuid_mod

def export_terrain_to_ifc(matrix_x, matrix_y, matrix_z, filepath, name="Terrain"):
    """
    Xuất lưới địa hình ra IFC tương thích Revit 2023+ (schema IFC4).
    Tọa độ VN-2000 được căn giữa về gốc; vị trí thực ghi vào ObjectPlacement của IfcSite.
    """
    try:
        import ifcopenshell
        from ifcopenshell import guid as ifc_guid
    except ImportError:
        st.error("Thiếu thư viện ifcopenshell. Hãy cài: pip install ifcopenshell")
        return False

    if matrix_x is None or matrix_y is None or matrix_z is None:
        st.error("Chưa có dữ liệu địa hình.")
        return False

    rows, cols = matrix_x.shape

    # Tọa độ VN-2000 thường rất lớn (> 500 000 m) → Revit không chấp nhận
    # → trừ offset về gốc, ghi offset vào ObjectPlacement của IfcSite
    x_off = float(np.nanmean(matrix_x))
    y_off = float(np.nanmean(matrix_y))

    # Xây dựng danh sách đỉnh (key = (i,j) để tránh lỗi float hash)
    idx_grid = {}   # (i,j) → 0-based index
    vertices  = []  # [[x-x_off, y-y_off, z], ...]

    for i in range(rows):
        for j in range(cols):
            xv = float(matrix_x[i, j])
            yv = float(matrix_y[i, j])
            zv = float(matrix_z[i, j])
            if np.isnan(xv) or np.isnan(yv) or np.isnan(zv):
                continue
            idx_grid[(i, j)] = len(vertices)
            vertices.append([xv - x_off, yv - y_off, zv])

    # Tạo tam giác (IFC dùng 1-based index)
    triangles = []
    for i in range(rows - 1):
        for j in range(cols - 1):
            corners = [(i,j), (i+1,j), (i,j+1), (i+1,j+1)]
            if not all(c in idx_grid for c in corners):
                continue
            v0, v1, v2, v3 = [idx_grid[c] for c in corners]

            def _tri(a, b, c):
                pa, pb, pc = np.array(vertices[a]), np.array(vertices[b]), np.array(vertices[c])
                normal = np.cross(pb - pa, pc - pa)
                return [a+1, b+1, c+1] if normal[2] >= 0 else [a+1, c+1, b+1]

            triangles.append(_tri(v0, v1, v2))
            triangles.append(_tri(v1, v3, v2))

    if not vertices or not triangles:
        st.error("Không đủ dữ liệu để xuất IFC (cần ít nhất 3 điểm hợp lệ).")
        return False

    # ── Tạo file IFC4 ──────────────────────────────────────────────────────
    f = ifcopenshell.file(schema="IFC4")
    try:
        f.wrapped_data.header.file_name.name = str(filepath)
        f.wrapped_data.header.file_name.author = ["UTH AI Bridge Design"]
        f.wrapped_data.header.file_name.organization = ["UTH"]
        f.wrapped_data.header.file_description.description = [
            "ViewDefinition [CoordinationView_V2.0]"
        ]
    except Exception:
        pass

    # OwnerHistory — Revit 2023+ từ chối file thiếu mục này
    person  = f.create_entity("IfcPerson",           Identification=None, FamilyName="UTH")
    org     = f.create_entity("IfcOrganization",     Identification=None, Name="UTH")
    p_org   = f.create_entity("IfcPersonAndOrganization", ThePerson=person, TheOrganization=org)
    app     = f.create_entity("IfcApplication",
                               ApplicationDeveloper=org, Version="1.0",
                               ApplicationFullName="AI Bridge Design",
                               ApplicationIdentifier="UTH-BIM")
    oh = f.create_entity("IfcOwnerHistory",
                          OwningUser=p_org, OwningApplication=app,
                          ChangeAction="ADDED", CreationDate=0)

    # Đơn vị
    units = f.create_entity("IfcUnitAssignment", Units=[
        f.create_entity("IfcSIUnit", UnitType="LENGTHUNIT",     Name="METRE"),
        f.create_entity("IfcSIUnit", UnitType="AREAUNIT",       Name="SQUARE_METRE"),
        f.create_entity("IfcSIUnit", UnitType="VOLUMEUNIT",     Name="CUBIC_METRE"),
        f.create_entity("IfcSIUnit", UnitType="PLANEANGLEUNIT", Name="RADIAN"),
    ])

    # Geometric context
    origin  = f.create_entity("IfcCartesianPoint", Coordinates=[0.0, 0.0, 0.0])
    z_dir   = f.create_entity("IfcDirection",      DirectionRatios=[0.0, 0.0, 1.0])
    x_dir   = f.create_entity("IfcDirection",      DirectionRatios=[1.0, 0.0, 0.0])
    wcs     = f.create_entity("IfcAxis2Placement3D", Location=origin, Axis=z_dir, RefDirection=x_dir)
    geo_ctx = f.create_entity("IfcGeometricRepresentationContext",
                               ContextType="Model",
                               CoordinateSpaceDimension=3,
                               Precision=1.0e-5,
                               WorldCoordinateSystem=wcs)
    body_ctx = f.create_entity("IfcGeometricRepresentationSubContext",
                                ContextIdentifier="Body",
                                ContextType="Model",
                                ParentContext=geo_ctx,
                                TargetView="MODEL_VIEW")

    # IfcProject — phải có RepresentationContexts và UnitsInContext
    proj = f.create_entity("IfcProject",
                           GlobalId=ifc_guid.compress(_uuid_mod.uuid4().hex),
                           OwnerHistory=oh,
                           Name="Bridge Terrain",
                           RepresentationContexts=[geo_ctx],
                           UnitsInContext=units)

    # Placement tại GỐC (toạ độ đã căn giữa). KHÔNG đặt site ở VN-2000 tuyệt đối
    # (≈590km) — sẽ nằm xa gốc khiến Revit không hiển thị.
    def _id_axis():
        return f.create_entity(
            "IfcAxis2Placement3D",
            Location=f.create_entity("IfcCartesianPoint", Coordinates=[0.0, 0.0, 0.0]),
            Axis=f.create_entity("IfcDirection", DirectionRatios=[0.0, 0.0, 1.0]),
            RefDirection=f.create_entity("IfcDirection", DirectionRatios=[1.0, 0.0, 0.0]))
    site_pl = f.create_entity("IfcLocalPlacement", RelativePlacement=_id_axis())

    # Địa hình (IfcTriangulatedFaceSet)
    pt_list = f.create_entity("IfcCartesianPointList3D", CoordList=vertices)
    tfs     = f.create_entity("IfcTriangulatedFaceSet",
                               Coordinates=pt_list,
                               CoordIndex=triangles,
                               Closed=False)
    shape_rep = f.create_entity("IfcShapeRepresentation",
                                 ContextOfItems=body_ctx,
                                 RepresentationIdentifier="Body",
                                 RepresentationType="Tessellation",
                                 Items=[tfs])
    prod_shape = f.create_entity("IfcProductDefinitionShape",
                                  Representations=[shape_rep])

    # IfcSite = vùng chứa (KHÔNG giữ shape — Revit thường không hiện geometry gắn
    # thẳng IfcSite). Lưới đưa vào IfcBuildingElementProxy → Revit luôn hiển thị.
    site = f.create_entity("IfcSite",
                           GlobalId=ifc_guid.compress(_uuid_mod.uuid4().hex),
                           OwnerHistory=oh, Name="Site",
                           ObjectPlacement=site_pl, CompositionType="ELEMENT")
    f.create_entity("IfcRelAggregates",
                    GlobalId=ifc_guid.compress(_uuid_mod.uuid4().hex),
                    OwnerHistory=oh, RelatingObject=proj, RelatedObjects=[site])

    terr_pl = f.create_entity("IfcLocalPlacement",
                              PlacementRelTo=site_pl, RelativePlacement=_id_axis())
    terr = f.create_entity("IfcBuildingElementProxy",
                           GlobalId=ifc_guid.compress(_uuid_mod.uuid4().hex),
                           OwnerHistory=oh, Name=name,
                           ObjectPlacement=terr_pl, Representation=prod_shape)
    f.create_entity("IfcRelContainedInSpatialStructure",
                    GlobalId=ifc_guid.compress(_uuid_mod.uuid4().hex),
                    OwnerHistory=oh, RelatingStructure=site, RelatedElements=[terr])

    f.write(filepath)
    return True