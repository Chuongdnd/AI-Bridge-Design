"""
Module 12 — Dự toán Sơ bộ Tự động (TKCS)
==========================================
Căn cứ: QĐ 409/2025/QĐ-BXD về suất đầu tư xây dựng công trình;
         TCVN 11823:2017; TCVN 10304:2014.

Hàm xuất công khai:
  tinh_du_toan_so_bo(res, **kwargs)         → dict  (dự toán đầy đủ)
  so_sanh_du_toan_3_pa(res)                 → dict  (so sánh 3 phương án)
  xuat_excel(du_toan_result, filepath)      → bool
  format_vnd(so_tien_nghin, unit="ty")      → str
"""

from __future__ import annotations
import math
from typing import Any

# ─────────────────────────────────────────────────────────────────────────────
# CƠ SỞ DỮ LIỆU ĐƠN GIÁ — tham chiếu QĐ 409/2025/QĐ-BXD
# Đơn vị: nghìn đồng (VND × 10³)
# ─────────────────────────────────────────────────────────────────────────────

DON_GIA_VAT_LIEU: dict[str, dict] = {
    # Bê tông tươi (vật liệu + phụ gia, chưa nhân công thi công)
    "BT_M30":  {"don_vi": "m³",  "don_gia": 1_500,  "mo_ta": "Bê tông thương phẩm M30"},
    "BT_M35":  {"don_vi": "m³",  "don_gia": 1_650,  "mo_ta": "Bê tông thương phẩm M35"},
    "BT_M40":  {"don_vi": "m³",  "don_gia": 1_850,  "mo_ta": "Bê tông thương phẩm M40"},
    "BT_M50":  {"don_vi": "m³",  "don_gia": 2_200,  "mo_ta": "Bê tông C50/60 dầm DUL"},
    "BT_coc":  {"don_vi": "m³",  "don_gia": 1_750,  "mo_ta": "Bê tông cọc khoan nhồi M30+"},
    # Cốt thép
    "CT_CB300": {"don_vi": "tấn", "don_gia": 16_500, "mo_ta": "Thép CB300-V phụ"},
    "CT_CB400": {"don_vi": "tấn", "don_gia": 18_500, "mo_ta": "Thép CB400-V chính"},
    "CT_DUL":   {"don_vi": "tấn", "don_gia": 36_000, "mo_ta": "Cáp DUL cường độ cao Φ12.7"},
    "CT_coc":   {"don_vi": "tấn", "don_gia": 18_500, "mo_ta": "Cốt thép cọc CB400"},
    # Cốp pha (giá khấu hao 1 lần sử dụng, chia cho ~15 lần tái dùng)
    "COPPHA_THEP": {"don_vi": "m²", "don_gia": 120,  "mo_ta": "Cốp pha thép — chi phí khấu hao + bảo dưỡng"},
    # Phụ trợ
    "GOI_CAU":    {"don_vi": "cái", "don_gia": 5_500,   "mo_ta": "Gối cao su bản thép (elastomeric)"},
    "KHE_CO_GIAN":{"don_vi": "m",   "don_gia": 4_500,   "mo_ta": "Khe co giãn (modular/finger)"},
    "LAN_CAN":    {"don_vi": "m",   "don_gia": 3_500,   "mo_ta": "Lan can bê tông + tay vịn thép"},
    "THOAT_NUOC": {"don_vi": "cái", "don_gia": 2_500,   "mo_ta": "Hệ thống thoát nước mặt cầu"},
    # Lớp phủ mặt cầu
    "ASPHALT":    {"don_vi": "m²",  "don_gia": 280,    "mo_ta": "Bê tông asphalt 70mm"},
    "LOP_PHONG_NUOC":{"don_vi": "m²","don_gia": 85,    "mo_ta": "Lớp phòng nước"},
}

DON_GIA_NHAN_CONG: dict[str, float] = {
    # Đơn vị: nghìn đồng/m³ (quy về bê tông tương đương)
    "duc_dam_DUL" : 850,   # Đúc dầm DUL tại xưởng (nâng hạ, căng cáp, bảo dưỡng)
    "lap_dam"     : 650,   # Lắp đặt dầm bằng cẩu
    "duc_tru"     : 420,   # Đổ bê tông trụ cầu
    "duc_mo"      : 380,   # Đổ bê tông mố cầu
    "duc_be_coc"  : 350,   # Đổ bê tông bệ cọc
    "khoan_nhoi"  : 2_800, # Thi công cọc khoan nhồi (nghìn đ/m dài cọc)
    "ep_coc"      : 850,   # Thi công cọc ép (nghìn đ/m dài cọc)
    "ban_mat_cau" : 380,   # Đổ bê tông bản mặt cầu
    "lap_thep"    : 1_800, # Lắp đặt cốt thép (nghìn đ/tấn)
    "lap_DUL"     : 4_200, # Lắp đặt + kéo căng cáp DUL (nghìn đ/tấn)
}

DON_GIA_MAY: dict[str, float] = {
    # Ca máy quy về nghìn đ/m³ bê tông hoặc đơn vị thích hợp
    "cau_lap_dam"  : 1_200,  # Cẩu lắp dầm (nghìn đ/m³ dầm)
    "may_bom_BT"   : 120,    # Máy bơm bê tông
    "may_tron_BT"  : 80,     # Máy trộn bê tông
    "may_khoan"    : 1_500,  # Máy khoan cọc nhồi (nghìn đ/m dài cọc)
    "may_ep_coc"   : 550,    # Máy ép cọc (nghìn đ/m dài cọc)
    "thiet_bi_DUL" : 800,    # Kích DUL + bơm + phụ kiện (nghìn đ/tấn cáp)
    "may_dam_BT"   : 45,     # Đầm dùi + đầm mặt
}

HE_SO_DIA_PHUONG: dict[str, float] = {
    "Hà Nội / TP.HCM"    : 1.00,
    "Đồng bằng Bắc Bộ"   : 0.92,
    "Đồng bằng Nam Bộ"   : 0.90,
    "Duyên hải miền Trung": 0.95,
    "Tây Nguyên"          : 1.12,
    "Trung du miền núi BB": 1.18,
    "Miền núi phía Bắc"   : 1.25,
    "Vùng sâu vùng xa"    : 1.38,
}

HE_SO_CAP_DUONG: dict[str, float] = {
    "Cao tốc"  : 1.18,
    "Cấp I"    : 1.12,
    "Cấp II"   : 1.07,
    "Cấp III"  : 1.00,
    "Cấp IV"   : 0.94,
    "Cấp V"    : 0.87,
    "Cấp VI"   : 0.82,
}

# Tỉ lệ cốt thép kinh nghiệm (kg/m³ bê tông dầm)
_CT_RATIO: dict[str, float] = {
    "Super-T"       : 145,
    "Dầm I"         : 130,
    "T ngược"       : 105,
    "Dầm T"         : 110,
    "Dầm bản rỗng"  : 95,
    "_default"      : 120,
}

_MM2M = 0.001


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def format_vnd(so_tien_nghin: float, unit: str = "ty") -> str:
    """Định dạng số tiền (đơn vị nghìn đồng) thành chuỗi dễ đọc."""
    if unit == "ty":
        return f"{so_tien_nghin / 1_000_000:.3f} tỷ"
    elif unit == "trieu":
        return f"{so_tien_nghin / 1_000:.1f} triệu"
    elif unit == "nghin":
        return f"{so_tien_nghin:,.0f} nghìn"
    return str(so_tien_nghin)


def _hang_muc(ten: str, don_vi: str, kl: float, don_gia: float,
              ghi_chu: str = "") -> dict:
    """Tạo dict hạng mục dự toán chuẩn hóa."""
    return {
        "ten"        : ten,
        "don_vi"     : don_vi,
        "khoi_luong" : round(kl, 3),
        "don_gia"    : round(don_gia, 0),
        "thanh_tien" : round(kl * don_gia, 0),   # nghìn đồng
        "ghi_chu"    : ghi_chu,
    }


def _poly_area(pts: list[tuple]) -> float:
    """Diện tích đa giác (shoelace formula). pts là danh sách (x,y)."""
    n = len(pts)
    if n < 3:
        return 0.0
    area = 0.0
    for i in range(n):
        x1, y1 = pts[i]
        x2, y2 = pts[(i + 1) % n]
        area += x1 * y2 - x2 * y1
    return abs(area) / 2.0


def _tiet_dien_mm2(beam_params: dict, loai_dam: str) -> float:
    """Diện tích tiết diện ngang dầm (mm²)."""
    mc = beam_params.get("MC_AA", {})
    H  = beam_params.get("H", 1750)

    try:
        from shapely.geometry import Polygon as ShPoly
        from shapely.affinity import scale as sh_scale
        import importlib.util, os
        _dir = os.path.dirname(os.path.abspath(__file__))
        spec = importlib.util.spec_from_file_location("beam3d", os.path.join(_dir, "04b-Beam3D.py"))
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        poly = mod.beam_section_polygon(beam_params, loai_dam)
        if poly is not None:
            return poly.area    # mm²
    except Exception:
        pass

    # Fallback: hình học đơn giản
    # Hỗ trợ cả 2 quy ước đặt tên key: Vietnamese (B_canh_tren…) và bf_top/tf_top
    def _get(vn_key, en_key, default):
        return float(mc.get(vn_key) or mc.get(en_key) or default)

    B = float(beam_params.get("B") or mc.get("B_canh_tren") or mc.get("bf_top") or 2440)
    if loai_dam == "Super-T":
        Bct  = _get("B_canh_tren",   "bf_top",  2440)
        Hct  = _get("H_canh_tren",   "tf_top",  200)
        bw   = _get("B_bung_top",    "bw",      200)
        Hbung= _get("H_bung",        "hw",      H - 2*200 - 150)
        Bcd  = _get("B_canh_duoi",   "bf_bot",  700)
        Hcd  = _get("H_canh_duoi",   "tf_bot",  200)
        # I-beam approximation: top flange + web + bottom flange
        A_gross = Bct * Hct + bw * Hbung + Bcd * Hcd
        # Super-T có khoang rỗng và vát cánh → giảm ~8%
        return A_gross * 0.92
    elif loai_dam in ("Dầm I", "T ngược", "Dầm T"):
        Bct  = _get("B_canh_tren", "bf_top",  600)
        Hct  = _get("H_canh_tren", "tf_top",  130)
        Bcd  = _get("B_canh_duoi", "bf_bot",  500)
        Hcd  = _get("H_canh_duoi", "tf_bot",  130)
        bw_  = float(mc.get("B_bung") or mc.get("bw") or 200)
        Hbung = H - Hct - Hcd
        return Bct * Hct + bw_ * Hbung + Bcd * Hcd
    elif loai_dam == "Dầm bản rỗng":
        B_rong = float(mc.get("B_rong", 300))
        n_rong = int(mc.get("so_rong", 3))
        A_void = n_rong * math.pi * (B_rong / 2) ** 2
        return B * H - A_void
    return B * H * 0.65   # generic fallback


# =============================================================================
# PHẦN 2 — TÍNH KHỐI LƯỢNG DẦM KẾT CẤU NHỊP
# =============================================================================

def tinh_khoi_luong_dam(beam_params: dict, n_dam: int, n_nhip: int = 1) -> dict:
    """
    Tính khối lượng vật liệu kết cấu nhịp (dầm).

    Parameters
    ----------
    beam_params : dict — beam_params_final từ Module 06b
    n_dam       : int  — số dầm trên 1 nhịp
    n_nhip      : int  — số nhịp

    Returns
    -------
    dict với các key: bt_m3, thep_thuong_tan, thep_DUL_tan, coppha_m2,
                      n_dam_tong, L_dam_m, ghi_chu
    """
    mc     = beam_params.get("MC_AA", {})
    loai   = beam_params.get("loai_dam", "Super-T")
    L_mm   = beam_params.get("L", 38200)
    L_m    = L_mm * _MM2M
    L_dau  = beam_params.get("MC_BB", {}).get("L_dau_dac", 750) * _MM2M
    cap    = beam_params.get("cap_DUL", {})
    n_dam_tong = n_dam * n_nhip

    # ── Diện tích tiết diện ──
    A_mid_mm2  = _tiet_dien_mm2(beam_params, loai)            # giữa nhịp
    # Tiết diện đầu dầm = đặc hoàn toàn (khoang rỗng bịt kín)
    A_end_mm2  = _tiet_dien_mm2({**beam_params,
                                  "MC_AA": {**mc, "B_rong_tren": 0, "B_rong_duoi": 0}}, loai) \
                 if loai in ("Super-T", "Dầm bản rỗng") else A_mid_mm2

    A_mid_m2 = A_mid_mm2 * _MM2M * _MM2M
    A_end_m2 = A_end_mm2 * _MM2M * _MM2M

    L_mid = max(L_m - 2 * L_dau, 0.5)
    vol_1_dam = A_end_m2 * 2 * L_dau + A_mid_m2 * L_mid   # m³/dầm
    bt_m3_tong = vol_1_dam * n_dam_tong

    # ── Cốt thép thường ──
    ratio_kg_m3 = _CT_RATIO.get(loai, _CT_RATIO["_default"])
    thep_thuong_tan = bt_m3_tong * ratio_kg_m3 / 1000.0

    # ── Cốt thép DUL ──
    so_ong    = cap.get("so_ong", 0)
    so_tao    = cap.get("so_tao_cap", 12)   # tao cáp trên 1 ống (mặc định 12 tao Ø12.7)
    kl_1tao_kgm = 0.775                      # kg/m, strand Ø12.7 mm
    L_cap     = L_m * 1.04                   # thêm 4% cho đầu neo
    thep_DUL_ton_1dam = so_ong * so_tao * L_cap * kl_1tao_kgm / 1000.0
    thep_DUL_tan = thep_DUL_ton_1dam * n_dam_tong

    # ── Cốp pha ──
    # Chu vi tiết diện xấp xỉ = 2×(B_outer + H)
    B_out   = beam_params.get("B", mc.get("B_canh_tren", 2440)) * _MM2M
    H_m     = beam_params.get("H", 1750) * _MM2M
    chu_vi  = 2 * (B_out + H_m)   # m (ước tính đơn giản)
    coppha_m2 = chu_vi * L_m * n_dam_tong

    return {
        "loai_dam"       : loai,
        "n_dam_tong"     : n_dam_tong,
        "L_dam_m"        : round(L_m, 2),
        "A_tiet_dien_m2" : round(A_mid_m2, 4),
        "bt_m3"          : round(bt_m3_tong, 2),
        "thep_thuong_tan": round(thep_thuong_tan, 3),
        "thep_DUL_tan"   : round(thep_DUL_tan, 3),
        "coppha_m2"      : round(coppha_m2, 1),
        "ghi_chu"        : (f"{n_dam_tong} dầm × {L_m:.1f}m; "
                            f"A={A_mid_m2:.4f}m²; "
                            f"V/dầm={vol_1_dam:.2f}m³"),
    }


# =============================================================================
# PHẦN 3 — TÍNH KHỐI LƯỢNG MỐ TRỤ
# =============================================================================

def tinh_khoi_luong_mo_tru(res: dict) -> dict:
    """
    Tính khối lượng vật liệu mố và trụ cầu.

    Đọc từ res: H_tru_est, tru_result, bc, geo_logic.
    """
    tru  = res.get("tru_result") or {}
    kcn  = res.get("kcn_result") or res.get("ai_result") or {}
    bc   = float(res.get("bc", 12.0))
    H_tru = float(res.get("H_tru_est", 6.0))
    n_nhip = int(kcn.get("tong_so_nhip", 1))
    n_tru  = max(0, n_nhip - 1)     # số trụ giữa nhịp

    loai_tru = str(tru.get("loai_tru", "Thân đặc"))
    W_tru = max(1.2, bc * 0.08)     # chiều dày thân trụ (m)
    L_tru = max(2.5, bc * 0.25)     # chiều dài xà mũ (m)

    # ── Thân trụ ──
    if "2 cột" in loai_tru.lower() or "2cot" in loai_tru.lower():
        D_col  = max(1.0, bc * 0.07)
        A_col  = math.pi * (D_col / 2) ** 2
        vol_than_1tru = 2 * A_col * H_tru
    elif "rỗng" in loai_tru.lower() or "rong" in loai_tru.lower():
        tWall   = max(0.3, W_tru * 0.18)
        A_out   = W_tru * L_tru
        A_in    = (W_tru - 2 * tWall) * (L_tru - 2 * tWall)
        vol_than_1tru = max(A_out - A_in, A_out * 0.4) * H_tru
    else:  # Thân đặc
        vol_than_1tru = W_tru * L_tru * H_tru

    # Xà mũ
    cap_H    = max(0.8, bc * 0.06)
    cap_L    = min(bc * 0.9, L_tru + 1.2)
    vol_xa_mu_1tru = cap_L * (W_tru + 0.5) * cap_H

    # Bệ trụ (phần trên đất)
    be_H   = 1.5
    be_L   = cap_L + 0.5
    be_W   = W_tru + 1.0
    vol_be_1tru = be_L * be_W * be_H

    vol_tru_tong = (vol_than_1tru + vol_xa_mu_1tru + vol_be_1tru) * n_tru

    # ── Mố cầu (2 mố, loại chữ U hoặc chân dê) ──
    # Xà mũ mố
    xa_mu_mo_vol = cap_L * (W_tru + 0.5) * cap_H * 2
    # Tường thân mố
    H_mo  = max(H_tru * 0.3, 2.0)
    than_mo_vol = 2 * (cap_L * 0.8 * H_mo)
    # Tường cánh mố
    L_tuong_canh = max(2.0, H_mo * 1.2)
    canh_mo_vol  = 4 * (L_tuong_canh * 0.5 * H_mo)

    vol_mo_tong = xa_mu_mo_vol + than_mo_vol + canh_mo_vol

    # ── Tổng khối lượng thép ──
    vol_BT_tong = vol_tru_tong + vol_mo_tong
    # Trụ + mố dùng M35 (thường), hệ số cốt thép 110 kg/m³
    thep_mo_tru_tan = vol_BT_tong * 110 / 1000.0

    return {
        "n_tru"            : n_tru,
        "loai_tru"         : loai_tru,
        "vol_than_tru_m3"  : round(vol_than_1tru * n_tru, 2),
        "vol_xa_mu_tru_m3" : round(vol_xa_mu_1tru * n_tru, 2),
        "vol_be_tru_m3"    : round(vol_be_1tru * n_tru, 2),
        "vol_mo_m3"        : round(vol_mo_tong, 2),
        "vol_BT_M35_m3"    : round(vol_BT_tong, 2),
        "thep_mo_tru_tan"  : round(thep_mo_tru_tan, 3),
        "ghi_chu"          : f"{n_tru} trụ {loai_tru}; 2 mố",
    }


# =============================================================================
# PHẦN 3B — TÍNH KHỐI LƯỢNG MÓNG CỌC
# =============================================================================

def tinh_khoi_luong_mong_coc(res: dict, n_tru: int) -> dict:
    """
    Tính khối lượng móng cọc (cọc + bệ móng).
    n_tru : số trụ giữa nhịp (mố luôn = 2)
    """
    mong = res.get("mong_result") or {}
    kcn  = res.get("kcn_result") or res.get("ai_result") or {}
    bc   = float(res.get("bc", 12.0))

    loai_coc   = str(mong.get("loai_mong", "Cọc khoan nhồi"))
    D_coc_mm   = float(mong.get("D_coc_mm", mong.get("kich_thuoc_mm", 800)))
    D_m        = D_coc_mm * _MM2M
    L_coc_m    = float(mong.get("chieu_dai_coc", mong.get("L_coc_tu", 20)))
    so_coc_be  = int(mong.get("so_coc_be", mong.get("So_coc_tu", 5)))
    Be_ngang   = float(mong.get("Be_ngang", max(3.0, D_m * 4)))
    Be_doc     = float(mong.get("Be_doc",   max(3.0, D_m * 4)))
    Be_cao     = float(mong.get("Be_cao",   max(1.0, D_m * 1.5)))

    n_be_tong  = n_tru + 2   # tổng số bệ móng (trụ + 2 mố)

    # Tổng chiều dài cọc
    L_coc_tong = L_coc_m * so_coc_be * n_be_tong

    # Thể tích bê tông cọc
    if "khoan nhồi" in loai_coc.lower():
        A_coc_m2 = math.pi * (D_m / 2) ** 2
        vol_coc_m3 = A_coc_m2 * L_coc_tong
    else:
        A_coc_m2 = (D_m ** 2)    # vuông
        vol_coc_m3 = A_coc_m2 * L_coc_tong

    # Thể tích bê tông bệ móng
    vol_be_m3 = Be_ngang * Be_doc * Be_cao * n_be_tong

    # Cốt thép cọc: 85 kg/m³
    thep_coc_tan = vol_coc_m3 * 85 / 1000.0
    # Cốt thép bệ: 100 kg/m³
    thep_be_tan  = vol_be_m3 * 100 / 1000.0

    return {
        "loai_coc"       : loai_coc,
        "D_coc_mm"       : D_coc_mm,
        "L_coc_m"        : L_coc_m,
        "so_coc_be"      : so_coc_be,
        "n_be_tong"      : n_be_tong,
        "L_coc_tong_m"   : round(L_coc_tong, 1),
        "vol_coc_m3"     : round(vol_coc_m3, 2),
        "vol_be_m3"      : round(vol_be_m3, 2),
        "thep_coc_tan"   : round(thep_coc_tan, 3),
        "thep_be_tan"    : round(thep_be_tan, 3),
        "ghi_chu"        : (f"{n_be_tong} bệ × {so_coc_be} cọc "
                            f"{loai_coc} Ø{D_coc_mm:.0f}mm L={L_coc_m}m"),
    }


# =============================================================================
# PHẦN 4 — HÀNG MỤC BẢN MẶT CẦU & PHỤ TRỢ
# =============================================================================

def tinh_khoi_luong_ban_mat_cau(res: dict) -> dict:
    """Bản mặt cầu BTCT M30."""
    kcn  = res.get("kcn_result") or res.get("ai_result") or {}
    bc   = float(res.get("bc", 12.0))
    L_cau = float((res.get("geo_logic") or {}).get("L_cau",
                   float(kcn.get("tong_so_nhip", 1)) * float(kcn.get("chieu_dai", 40))))
    t_ban = float(res.get("t_ban_mm", 200)) / 1000.0

    A_ban  = bc * L_cau                      # m²
    vol_bt = A_ban * t_ban                   # m³
    thep   = vol_bt * 110 / 1000.0           # tấn thép CB400

    lop_phu = res.get("lop_phu_result") or {}
    t_phu_mm = float(lop_phu.get("tong_day_tt", 70))
    A_asphalt = A_ban
    A_phong_nuoc = A_ban

    return {
        "A_ban_m2"      : round(A_ban, 1),
        "vol_BT_M30_m3" : round(vol_bt, 2),
        "thep_ban_tan"  : round(thep, 3),
        "A_asphalt_m2"  : round(A_asphalt, 1),
        "A_phong_nuoc_m2": round(A_phong_nuoc, 1),
        "t_phu_mm"      : t_phu_mm,
        "ghi_chu"       : f"Bản BTCT {t_ban*1000:.0f}mm + lớp phủ {t_phu_mm:.0f}mm",
    }


def tinh_hang_muc_phu_tro(res: dict, n_nhip: int) -> dict:
    """Gối cầu, khe co giãn, lan can, thoát nước."""
    kcn  = res.get("kcn_result") or res.get("ai_result") or {}
    bc   = float(res.get("bc", 12.0))
    n_dam = int(kcn.get("so_luong_dam") or kcn.get("so_luong_dam_mcn", 5))
    L_cau = float((res.get("geo_logic") or {}).get("L_cau",
                   n_nhip * float(kcn.get("chieu_dai", 40))))

    # Gối cầu: 1 gối/đầu dầm tại mỗi đường gối (mố + trụ)
    n_goi = n_dam * (n_nhip + 1)  # đường gối = nhịp + 1

    # Khe co giãn: 1 khe/nhịp ở 2 đầu + 1 khe/2 nhịp giữa
    n_khe   = (n_nhip + 1)
    L_khe_m = bc * n_khe  # tổng m khe co giãn

    # Lan can: 2 bên
    L_lan_can = L_cau * 2

    # Thoát nước mặt cầu: mỗi 6m một hố ga
    n_thoat = math.ceil(L_cau / 6) * 2

    return {
        "n_goi_cau"    : n_goi,
        "L_khe_co_gian": round(L_khe_m, 1),
        "L_lan_can_m"  : round(L_lan_can, 1),
        "n_thoat_nuoc" : n_thoat,
        "ghi_chu"       : f"{n_goi} gối; {L_khe_m:.0f}m khe; {L_lan_can:.0f}m lan can",
    }


# =============================================================================
# PHẦN 4 — DỰ TOÁN TỔNG HỢP CHÍNH
# =============================================================================

def tinh_du_toan_so_bo(
    res: dict,
    vung_mien: str  = "Đồng bằng Nam Bộ",
    cap_duong: str  = "Cấp III",
    he_so_dt_pct: float = 10.0,
    he_so_tk_ql: float  = 5.0,
    he_so_ks:    float  = 2.0,
    he_so_tn:    float  = 1.0,
    he_so_bh:    float  = 0.5,
    vat_pct:     float  = 10.0,
    beam_params: dict | None = None,
) -> dict:
    """
    Tính dự toán sơ bộ TKCS toàn công trình cầu.

    Parameters
    ----------
    res         : dict — st.session_state["res"] từ Module 00-Interface
    vung_mien   : str  — vùng miền áp dụng hệ số địa phương
    cap_duong   : str  — cấp đường để lấy hệ số kỹ thuật
    he_so_*     : float — tỉ lệ % chi phí phụ trợ
    beam_params : dict | None — nếu None, đọc từ res["beam_params_final"]

    Returns
    -------
    dict:
        hang_muc_1  (KCN - Kết cấu nhịp)
        hang_muc_2  (MTM - Mố trụ)
        hang_muc_3  (MON - Móng cọc)
        hang_muc_4  (BMC - Bản mặt cầu)
        hang_muc_5  (PHU - Phụ trợ)
        tong_hop    (tổng chi phí các cấp)
        thong_so    (thông số tổng quát)
    """
    DGV = DON_GIA_VAT_LIEU
    DGC = DON_GIA_NHAN_CONG
    DGM = DON_GIA_MAY

    k_dp  = HE_SO_DIA_PHUONG.get(vung_mien, 0.92)
    k_cap = HE_SO_CAP_DUONG.get(cap_duong,  1.00)
    k     = k_dp * k_cap    # hệ số tổng hợp

    if beam_params is None:
        beam_params = res.get("beam_params_final")

    kcn   = res.get("kcn_result") or res.get("ai_result") or {}
    n_nhip  = int(kcn.get("tong_so_nhip", 1))
    n_dam   = int(kcn.get("so_luong_dam") or kcn.get("so_luong_dam_mcn", 5))
    bc      = float(res.get("bc", 12.0))

    if beam_params is None:
        # Ước tính từ kcn_result
        loai_dam = str(kcn.get("loai_dam", "Super-T"))
        L_mm = float(kcn.get("chieu_dai", 38.2)) * 1000
        H_mm = float(kcn.get("chieu_cao_dam") or kcn.get("chieu_cao", 1.75)) * 1000
        beam_params = {
            "loai_dam": loai_dam, "L": L_mm, "H": H_mm,
            "B": 2440, "S": float(kcn.get("khoang_cach_dam", 2.2)) * 1000,
            "MC_AA": {"B_canh_tren": 2200, "H_canh_tren": 150, "H_vat_canh_tren": 75,
                      "H_bung": int(H_mm - 450), "B_bung_top": 110, "B_bung_bot": 160,
                      "B_canh_duoi": 1020, "H_canh_duoi": 200, "H_vat_canh_duoi": 100,
                      "B_rong_tren": 800, "B_rong_duoi": 700, "web_out_hw": 480},
            "MC_BB": {"L_dau_dac": 750},
            "cap_DUL": {"so_ong": 5, "so_tao_cap": 12, "ong_PVC_D": 49},
        }

    # ── Khối lượng từng phần ──
    kl_dam  = tinh_khoi_luong_dam(beam_params, n_dam, n_nhip)
    kl_tru  = tinh_khoi_luong_mo_tru(res)
    kl_coc  = tinh_khoi_luong_mong_coc(res, kl_tru["n_tru"])
    kl_ban  = tinh_khoi_luong_ban_mat_cau(res)
    kl_phu  = tinh_hang_muc_phu_tro(res, n_nhip)

    # ─────────────────────────────────────────────────────────────────────────
    # HẠNG MỤC 1 — KẾT CẤU NHỊP
    # ─────────────────────────────────────────────────────────────────────────
    dg_bt_dam = DGV["BT_M50"]["don_gia"] + DGC["duc_dam_DUL"] + DGM["cau_lap_dam"] + DGM["may_bom_BT"]
    dg_ct_tt  = DGV["CT_CB400"]["don_gia"] + DGC["lap_thep"]
    dg_ct_DUL = DGV["CT_DUL"]["don_gia"]   + DGC["lap_DUL"] + DGM["thiet_bi_DUL"]
    dg_lap    = DGC["lap_dam"] * kl_dam["bt_m3"]    # chi phí lắp dầm (tính trên volume)

    hm1 = [
        _hang_muc("Bê tông dầm M50/60 — đúc + vận chuyển + lắp",
                   "m³",  kl_dam["bt_m3"],            dg_bt_dam * k,
                   kl_dam["ghi_chu"]),
        _hang_muc("Cốt thép thường CB400 dầm",
                   "tấn", kl_dam["thep_thuong_tan"],   dg_ct_tt  * k),
        _hang_muc("Cáp DUL cường độ cao Ø12.7 — lắp đặt + kéo căng",
                   "tấn", kl_dam["thep_DUL_tan"],      dg_ct_DUL * k),
        _hang_muc("Cốp pha dầm (thép tái sử dụng)",
                   "m²",  kl_dam["coppha_m2"],
                   (DGV["COPPHA_THEP"]["don_gia"] + 50) * k),
        _hang_muc("Chi phí bệ đúc + kéo căng + bảo dưỡng dầm",
                   "dầm", kl_dam["n_dam_tong"],
                   3_500 * k,   "Theo kinh nghiệm thực tế công trường"),
    ]

    # ─────────────────────────────────────────────────────────────────────────
    # HẠNG MỤC 2 — MỐ TRỤ
    # ─────────────────────────────────────────────────────────────────────────
    dg_bt_tru = DGV["BT_M35"]["don_gia"] + DGC["duc_tru"] + DGM["may_bom_BT"] + DGM["may_dam_BT"]
    dg_bt_mo  = DGV["BT_M35"]["don_gia"] + DGC["duc_mo"]  + DGM["may_bom_BT"] + DGM["may_dam_BT"]
    vol_tru_only = kl_tru["vol_than_tru_m3"] + kl_tru["vol_xa_mu_tru_m3"] + kl_tru["vol_be_tru_m3"]
    vol_mo_only  = kl_tru["vol_mo_m3"]

    hm2 = [
        _hang_muc("Bê tông trụ M35 — thân + xà mũ + bệ trụ",
                   "m³",  vol_tru_only,              dg_bt_tru * k,
                   kl_tru["ghi_chu"]),
        _hang_muc("Bê tông mố cầu M35",
                   "m³",  vol_mo_only,               dg_bt_mo  * k),
        _hang_muc("Cốt thép mố trụ CB400",
                   "tấn", kl_tru["thep_mo_tru_tan"], dg_ct_tt  * k),
        _hang_muc("Cốp pha mố trụ (thép)",
                   "m²",  kl_tru["vol_BT_M35_m3"] * 5,
                   DGV["COPPHA_THEP"]["don_gia"] * k,
                   "Ước tính 5 m²/m³ bê tông"),
    ]

    # ─────────────────────────────────────────────────────────────────────────
    # HẠNG MỤC 3 — MÓNG CỌC
    # ─────────────────────────────────────────────────────────────────────────
    is_bored = "khoan nhồi" in kl_coc["loai_coc"].lower()
    dg_coc_m = (DGC["khoan_nhoi"] + DGM["may_khoan"]
                if is_bored
                else DGC["ep_coc"] + DGM["may_ep_coc"])
    dg_bt_be   = DGV["BT_M30"]["don_gia"] + DGC["duc_be_coc"] + DGM["may_bom_BT"]
    dg_coc_vl  = (DGV["BT_coc"]["don_gia"] + 80  # bê tông cọc nhồi (~80 nhân công/m³)
                  if is_bored
                  else DGV["BT_M35"]["don_gia"])
    # Đơn giá/m dài cọc = VL/m + nhân công + máy
    D_m = kl_coc["D_coc_mm"] * _MM2M
    A_coc_m2 = math.pi * (D_m / 2) ** 2 if is_bored else D_m ** 2
    dg_per_m  = dg_coc_vl * A_coc_m2 + dg_coc_m   # nghìn đ/m dài cọc

    hm3 = [
        _hang_muc(f"Thi công cọc {kl_coc['loai_coc']} Ø{kl_coc['D_coc_mm']:.0f}mm",
                   "m",   kl_coc["L_coc_tong_m"],   dg_per_m * k,
                   kl_coc["ghi_chu"]),
        _hang_muc("Bê tông bệ cọc M30",
                   "m³",  kl_coc["vol_be_m3"],       dg_bt_be * k),
        _hang_muc("Cốt thép bệ cọc CB400",
                   "tấn", kl_coc["thep_be_tan"],     dg_ct_tt * k),
        _hang_muc("Cốt thép trong cọc CB400",
                   "tấn", kl_coc["thep_coc_tan"],    DGV["CT_coc"]["don_gia"] * k),
        _hang_muc("Thí nghiệm kiểm tra cọc (PDA / nén tĩnh)",
                   "cọc", max(2, kl_coc["n_be_tong"]),
                   8_500 * k, "Tối thiểu 2 cọc thí nghiệm"),
    ]

    # ─────────────────────────────────────────────────────────────────────────
    # HẠNG MỤC 4 — BẢN MẶT CẦU + LỚP PHỦ
    # ─────────────────────────────────────────────────────────────────────────
    dg_bt_ban = DGV["BT_M30"]["don_gia"] + DGC["ban_mat_cau"] + DGM["may_bom_BT"]

    hm4 = [
        _hang_muc("Bê tông bản mặt cầu M30",
                   "m³",  kl_ban["vol_BT_M30_m3"], dg_bt_ban * k,
                   kl_ban["ghi_chu"]),
        _hang_muc("Cốt thép bản mặt cầu CB400",
                   "tấn", kl_ban["thep_ban_tan"],  dg_ct_tt * k),
        _hang_muc("Cốp pha bản mặt cầu",
                   "m²",  kl_ban["A_ban_m2"],
                   DGV["COPPHA_THEP"]["don_gia"] * k),
        _hang_muc("Lớp phòng nước bản mặt cầu",
                   "m²",  kl_ban["A_phong_nuoc_m2"],
                   DGV["LOP_PHONG_NUOC"]["don_gia"] * k),
        _hang_muc("Lớp phủ bê tông asphalt",
                   "m²",  kl_ban["A_asphalt_m2"],
                   DGV["ASPHALT"]["don_gia"] * k,
                   f"Dày {kl_ban['t_phu_mm']:.0f}mm"),
    ]

    # ─────────────────────────────────────────────────────────────────────────
    # HẠNG MỤC 5 — PHỤ TRỢ
    # ─────────────────────────────────────────────────────────────────────────
    hm5 = [
        _hang_muc("Gối cầu cao su bản thép",
                   "cái", kl_phu["n_goi_cau"],
                   DGV["GOI_CAU"]["don_gia"] * k,
                   kl_phu["ghi_chu"]),
        _hang_muc("Khe co giãn",
                   "m",   kl_phu["L_khe_co_gian"],
                   DGV["KHE_CO_GIAN"]["don_gia"] * k),
        _hang_muc("Lan can BTCT + tay vịn",
                   "m",   kl_phu["L_lan_can_m"],
                   DGV["LAN_CAN"]["don_gia"] * k),
        _hang_muc("Hệ thống thoát nước mặt cầu",
                   "cái", kl_phu["n_thoat_nuoc"],
                   DGV["THOAT_NUOC"]["don_gia"] * k),
        _hang_muc("Chiếu sáng, biển báo, sơn kẻ đường",
                   "m",   kl_ban["A_ban_m2"] / bc,
                   800 * k, "Ước tính theo chiều dài cầu"),
    ]

    # ─────────────────────────────────────────────────────────────────────────
    # TỔNG HỢP
    # ─────────────────────────────────────────────────────────────────────────
    def _tong(hm: list) -> float:
        return sum(h["thanh_tien"] for h in hm)

    t1, t2, t3, t4, t5 = _tong(hm1), _tong(hm2), _tong(hm3), _tong(hm4), _tong(hm5)
    tong_xay_lap = t1 + t2 + t3 + t4 + t5

    chi_phi_dt   = tong_xay_lap * he_so_dt_pct / 100
    chi_phi_tk   = tong_xay_lap * he_so_tk_ql  / 100
    chi_phi_ks   = tong_xay_lap * he_so_ks     / 100
    chi_phi_tn   = tong_xay_lap * he_so_tn     / 100
    chi_phi_bh   = tong_xay_lap * he_so_bh     / 100
    tong_truoc_VAT = tong_xay_lap + chi_phi_dt + chi_phi_tk + chi_phi_ks + chi_phi_tn + chi_phi_bh
    vat            = tong_truoc_VAT * vat_pct / 100
    tong_dau_tu    = tong_truoc_VAT + vat

    # Chỉ số phân tích
    L_cau = float((res.get("geo_logic") or {}).get(
        "L_cau", n_nhip * float(kcn.get("chieu_dai", 38.2))))
    A_cau_m2 = bc * L_cau
    suat_dau_tu = tong_dau_tu / A_cau_m2   # nghìn đ/m² → triệu/m² là /1000

    return {
        "hang_muc_1" : hm1,
        "hang_muc_2" : hm2,
        "hang_muc_3" : hm3,
        "hang_muc_4" : hm4,
        "hang_muc_5" : hm5,
        "khoi_luong" : {
            "dam"  : kl_dam,
            "tru"  : kl_tru,
            "coc"  : kl_coc,
            "ban"  : kl_ban,
            "phu"  : kl_phu,
        },
        "tong_hop" : {
            "hm1_KCN"          : round(t1, 0),
            "hm2_MTM"          : round(t2, 0),
            "hm3_MON"          : round(t3, 0),
            "hm4_BMC"          : round(t4, 0),
            "hm5_PHU"          : round(t5, 0),
            "tong_xay_lap"     : round(tong_xay_lap, 0),
            "du_phong"         : round(chi_phi_dt, 0),
            "thiet_ke_QLDA"    : round(chi_phi_tk, 0),
            "khao_sat"         : round(chi_phi_ks, 0),
            "thu_nghiem"       : round(chi_phi_tn, 0),
            "bao_hiem"         : round(chi_phi_bh, 0),
            "tong_truoc_VAT"   : round(tong_truoc_VAT, 0),
            "VAT"              : round(vat, 0),
            "tong_dau_tu"      : round(tong_dau_tu, 0),
        },
        "phan_tich" : {
            "A_cau_m2"         : round(A_cau_m2, 1),
            "L_cau_m"          : round(L_cau, 1),
            "suat_dau_tu_ty_m2": round(suat_dau_tu / 1_000_000, 3),   # tỷ/m²... cần triệu
            "suat_dau_tu_trieu_m2": round(suat_dau_tu / 1_000, 2),    # triệu VND/m²
            "ty_le_KCN_pct"    : round(t1 / tong_xay_lap * 100, 1),
            "ty_le_MTM_pct"    : round(t2 / tong_xay_lap * 100, 1),
            "ty_le_MON_pct"    : round(t3 / tong_xay_lap * 100, 1),
            "canh_bao"         : _canh_bao_suat_dau_tu(suat_dau_tu / 1_000, cap_duong),
        },
        "thong_so" : {
            "vung_mien"  : vung_mien,
            "cap_duong"  : cap_duong,
            "k_dia_phuong": k_dp,
            "k_cap_duong" : k_cap,
            "n_nhip"     : n_nhip,
            "n_dam"      : n_dam,
            "bc"         : bc,
            "loai_dam"   : beam_params.get("loai_dam", "—"),
        },
    }


def _canh_bao_suat_dau_tu(suat_trieu_m2: float, cap_duong: str) -> str:
    """Đánh giá suất đầu tư theo loại cầu."""
    # Dải tham chiếu (triệu VND/m²) theo cấp đường, QĐ 409/2025
    khoang: dict[str, tuple] = {
        "Cao tốc": (14, 22), "Cấp I": (12, 18), "Cấp II": (10, 16),
        "Cấp III": (8, 13),  "Cấp IV": (6, 11),  "Cấp V": (5, 9), "Cấp VI": (4, 7),
    }
    lo, hi = khoang.get(cap_duong, (7, 14))
    if suat_trieu_m2 < lo * 0.85:
        return f"⚠️ Suất đầu tư {suat_trieu_m2:.1f} triệu/m² thấp hơn mức tham chiếu ({lo}–{hi}) — kiểm tra lại thông số"
    elif suat_trieu_m2 > hi * 1.20:
        return f"⚠️ Suất đầu tư {suat_trieu_m2:.1f} triệu/m² cao hơn mức tham chiếu ({lo}–{hi}) — xem xét tối ưu phương án"
    return f"✅ Suất đầu tư {suat_trieu_m2:.1f} triệu/m² trong dải tham chiếu ({lo}–{hi} triệu/m²)"


# =============================================================================
# PHẦN 6 — SO SÁNH BA PHƯƠNG ÁN
# =============================================================================

def so_sanh_du_toan_3_pa(res: dict,
                          vung_mien: str = "Đồng bằng Nam Bộ",
                          cap_duong: str = "Cấp III") -> dict:
    """
    Tính dự toán cho 3 phương án kết cấu nhịp và so sánh.

    Phương án 1: Từ kcn_result (phương án chính đã chọn)
    Phương án 2: Loại dầm thay thế 1 (nếu có trong alternatives)
    Phương án 3: Loại dầm thay thế 2

    Returns dict {pa1, pa2, pa3, so_sanh}
    """
    kcn  = res.get("kcn_result") or res.get("ai_result") or {}
    alts = res.get("alternatives") or []

    def _res_variant(loai_dam_override: str | None) -> dict:
        if loai_dam_override is None:
            return res
        # Clone res và ghi đè loại dầm
        import copy
        r = copy.deepcopy(res)
        kcn_v = (r.get("kcn_result") or r.get("ai_result") or {}).copy()
        kcn_v["loai_dam"] = loai_dam_override
        r["kcn_result"]       = kcn_v
        r["beam_params_final"] = None   # Bắt tính lại từ kcn
        return r

    loai_1 = str(kcn.get("loai_dam", "Super-T"))
    alt_loais = [a.get("loai_dam") for a in alts if isinstance(a, dict) and a.get("loai_dam")
                 and a["loai_dam"] != loai_1]
    loai_2 = alt_loais[0] if len(alt_loais) > 0 else "Dầm I"
    loai_3 = alt_loais[1] if len(alt_loais) > 1 else "T ngược"

    pa1 = tinh_du_toan_so_bo(_res_variant(None),     vung_mien, cap_duong)
    pa2 = tinh_du_toan_so_bo(_res_variant(loai_2),   vung_mien, cap_duong)
    pa3 = tinh_du_toan_so_bo(_res_variant(loai_3),   vung_mien, cap_duong)

    def _key_metrics(pa: dict) -> dict:
        th = pa["tong_hop"]
        ph = pa["phan_tich"]
        return {
            "loai_dam"              : pa["thong_so"]["loai_dam"],
            "chi_phi_KCN_ty"        : round(th["hm1_KCN"] / 1e6, 3),
            "chi_phi_MTM_ty"        : round(th["hm2_MTM"] / 1e6, 3),
            "chi_phi_MON_ty"        : round(th["hm3_MON"] / 1e6, 3),
            "tong_xay_lap_ty"       : round(th["tong_xay_lap"] / 1e6, 3),
            "tong_dau_tu_ty"        : round(th["tong_dau_tu"] / 1e6, 3),
            "suat_dau_tu_trieu_m2"  : ph["suat_dau_tu_trieu_m2"],
            "ty_le_KCN_pct"         : ph["ty_le_KCN_pct"],
        }

    m1, m2, m3 = _key_metrics(pa1), _key_metrics(pa2), _key_metrics(pa3)
    tong_min = min(m1["tong_dau_tu_ty"], m2["tong_dau_tu_ty"], m3["tong_dau_tu_ty"])

    for m in [m1, m2, m3]:
        m["chenh_so_min_ty"]  = round(m["tong_dau_tu_ty"] - tong_min, 3)
        m["chenh_so_min_pct"] = round((m["tong_dau_tu_ty"] / tong_min - 1) * 100, 1) if tong_min else 0

    # Điểm tổng hợp 100: kinh tế 30 điểm
    def _diem_kinh_te(m: dict) -> float:
        if m["tong_dau_tu_ty"] == 0:
            return 0
        return 30.0 * tong_min / m["tong_dau_tu_ty"]

    m1["diem_kinh_te"] = round(_diem_kinh_te(m1), 1)
    m2["diem_kinh_te"] = round(_diem_kinh_te(m2), 1)
    m3["diem_kinh_te"] = round(_diem_kinh_te(m3), 1)

    return {"pa1": pa1, "pa2": pa2, "pa3": pa3,
            "so_sanh": {"pa1": m1, "pa2": m2, "pa3": m3}}


# =============================================================================
# PHẦN 7 — XUẤT EXCEL
# =============================================================================

def xuat_excel(du_toan: dict, filepath: str | None = None) -> bytes | bool:
    """
    Xuất bảng dự toán sang Excel (.xlsx) theo định dạng chuẩn ngành XD.

    Trả về bytes nếu filepath=None (dùng với st.download_button),
    hoặc True/False nếu filepath được chỉ định.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import io as _io
    except ImportError:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DU TOAN SO BO"

    # ── Style helpers ──
    def _style(cell, bold=False, color=None, align="center", bg=None, size=11):
        cell.font      = Font(name="Times New Roman", size=size, bold=bold)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if color:
            cell.font = Font(name="Times New Roman", size=size, bold=bold, color=color)

    thin = Side(style="thin")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _border(cell):
        cell.border = bdr

    # ── Tiêu đề ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "BẢNG DỰ TOÁN SƠ BỘ CÔNG TRÌNH CẦU"
    _style(ws["A1"], bold=True, size=14)

    thong_so = du_toan.get("thong_so", {})
    ws.merge_cells("A2:G2")
    ws["A2"] = (f"Loại dầm: {thong_so.get('loai_dam','—')}  |  "
                f"Bề rộng: {thong_so.get('bc','—')}m  |  {thong_so.get('n_nhip','—')} nhịp  |  "
                f"Vùng: {thong_so.get('vung_mien','—')}  |  Cấp đường: {thong_so.get('cap_duong','—')}")
    _style(ws["A2"], size=10)

    # ── Header bảng ──
    headers = ["STT", "Tên hạng mục", "Đơn vị", "Khối lượng", "Đơn giá (nghìn đ)", "Thành tiền (nghìn đ)", "Ghi chú"]
    col_ws  = [5, 40, 10, 12, 18, 22, 25]

    for i, (h, w) in enumerate(zip(headers, col_ws), 1):
        c = ws.cell(row=4, column=i, value=h)
        _style(c, bold=True, bg="4472C4", color="FFFFFF")
        _border(c)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30

    # ── Dữ liệu hạng mục ──
    HM_CFG = [
        ("I", "KẾT CẤU NHỊP (KCN)", "hang_muc_1", "BDD7E8"),
        ("II", "MỐ TRỤ CẦU (MTM)", "hang_muc_2", "E2EFDA"),
        ("III", "MÓNG CỌC (MON)", "hang_muc_3", "FFF2CC"),
        ("IV", "BẢN MẶT CẦU + LỚP PHỦ", "hang_muc_4", "FCE4D6"),
        ("V", "PHỤ TRỢ", "hang_muc_5", "F2F2F2"),
    ]

    row = 5
    stt_main = 0
    for roman, ten_nhom, key, bg in HM_CFG:
        stt_main += 1
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1,
                    value=f"{roman}. {ten_nhom}")
        _style(c, bold=True, bg=bg, align="left")
        for col_i in range(1, 8):
            _border(ws.cell(row=row, column=col_i))
        row += 1

        for i_hm, hm in enumerate(du_toan.get(key, []), 1):
            ws.cell(row=row, column=1, value=f"{roman}.{i_hm}")
            ws.cell(row=row, column=2, value=hm["ten"])
            ws.cell(row=row, column=3, value=hm["don_vi"])
            ws.cell(row=row, column=4, value=hm["khoi_luong"])
            ws.cell(row=row, column=5, value=hm["don_gia"])
            ws.cell(row=row, column=6, value=hm["thanh_tien"])
            ws.cell(row=row, column=7, value=hm.get("ghi_chu", ""))
            for col_i in range(1, 8):
                c = ws.cell(row=row, column=col_i)
                _style(c, align="left" if col_i in (2, 7) else "center")
                _border(c)
                if col_i in (4, 5, 6):
                    c.number_format = "#,##0"
            row += 1

    # ── Tổng hợp ──
    th = du_toan.get("tong_hop", {})
    tong_rows = [
        ("", "TỔNG CHI PHÍ XÂY LẮP",                  th.get("tong_xay_lap", 0)),
        ("",  f"  Dự phòng ({du_toan.get('thong_so', {}).get('cap_duong', 'Cấp III')}) 10%",
                                                         th.get("du_phong", 0)),
        ("",  "  Thiết kế + QLDA 5%",                   th.get("thiet_ke_QLDA", 0)),
        ("",  "  Khảo sát 2%",                           th.get("khao_sat", 0)),
        ("",  "  Thử nghiệm 1%",                         th.get("thu_nghiem", 0)),
        ("",  "  Bảo hiểm công trình 0.5%",              th.get("bao_hiem", 0)),
        ("",  "TỔNG TRƯỚC THUẾ",                         th.get("tong_truoc_VAT", 0)),
        ("",  "  VAT 10%",                               th.get("VAT", 0)),
        ("",  "TỔNG VỐN ĐẦU TƯ (có VAT)",               th.get("tong_dau_tu", 0)),
    ]
    bold_rows = {0, 6, 8}  # indices

    for idx, (stt, ten, tien) in enumerate(tong_rows):
        ws.cell(row=row, column=1, value=stt)
        c2 = ws.cell(row=row, column=2, value=ten)
        ws.merge_cells(f"B{row}:E{row}")
        c6 = ws.cell(row=row, column=6, value=tien)
        ws.cell(row=row, column=7, value="")
        is_bold = idx in bold_rows
        _style(c2, bold=is_bold, align="left",
               bg=("C9DAF8" if idx in (0, 6, 8) else None))
        _style(c6, bold=is_bold,
               bg=("C9DAF8" if idx in (0, 6, 8) else None))
        c6.number_format = "#,##0"
        for col_i in range(1, 8):
            _border(ws.cell(row=row, column=col_i))
        row += 1

    # ── Phân tích ──
    ph = du_toan.get("phan_tich", {})
    row += 1
    ws.cell(row=row, column=1, value="Phân tích:")
    ws.cell(row=row, column=2, value=ph.get("canh_bao", ""))
    ws.merge_cells(f"B{row}:G{row}")
    _style(ws.cell(row=row, column=2), align="left", bold=True)
    row += 1
    ws.cell(row=row, column=2,
            value=f"Suất đầu tư: {ph.get('suat_dau_tu_trieu_m2', 0):.1f} triệu đồng/m² mặt cầu  "
                  f"| KCN {ph.get('ty_le_KCN_pct', 0):.0f}%  "
                  f"| MTM {ph.get('ty_le_MTM_pct', 0):.0f}%  "
                  f"| Móng {ph.get('ty_le_MON_pct', 0):.0f}%")
    ws.merge_cells(f"B{row}:G{row}")
    _style(ws.cell(row=row, column=2), align="left")

    ws.freeze_panes = "A5"

    buf = _io.BytesIO()
    wb.save(buf)
    if filepath:
        with open(filepath, "wb") as f:
            f.write(buf.getvalue())
        return True
    return buf.getvalue()


# =============================================================================
# CHẠY THỬ ĐỘC LẬP
# =============================================================================
if __name__ == "__main__":
    import sys, json
    sys.stdout.reconfigure(encoding="utf-8")

    # Dữ liệu mẫu: Super-T 38.2m, 1 nhịp, 12m rộng, đồng bằng
    sample_res = {
        "bc"         : 12.0,
        "vtk"        : 80,
        "H_tru_est"  : 5.0,
        "t_ban_mm"   : 200,
        "geo_logic"  : {"L_cau": 38.2},
        "ai_result"  : {
            "loai_dam"        : "Super-T",
            "tong_so_nhip"    : 1,
            "chieu_dai"       : 38.2,
            "chieu_cao_dam"   : 1.75,
            "so_luong_dam"    : 6,
            "khoang_cach_dam" : 2.0,
            "overhang"        : 1.0,
        },
        "kcn_result" : None,
        "tru_result" : {"loai_tru": "Thân đặc"},
        "mong_result": {
            "loai_mong"   : "Cọc khoan nhồi",
            "D_coc_mm"    : 800,
            "chieu_dai_coc": 20.0,
            "so_coc_be"   : 5,
            "Be_ngang"    : 3.5,
            "Be_doc"      : 3.5,
            "Be_cao"      : 1.5,
        },
        "lop_phu_result": {"tong_day_tt": 70},
    }

    sample_beam = {
        "loai_dam": "Super-T", "L": 38200, "H": 1750, "B": 2440, "S": 2000,
        "MC_AA": {
            "B_canh_tren": 2200, "H_canh_tren": 150, "H_vat_canh_tren": 75,
            "B_vat_canh_tren": 100, "H_bung": 1250,
            "B_bung_top": 110, "B_bung_bot": 160,
            "B_canh_duoi": 1020, "H_canh_duoi": 200, "H_vat_canh_duoi": 100,
            "B_rong_tren": 800, "B_rong_duoi": 700, "web_out_hw": 480,
        },
        "MC_BB": {"L_dau_dac": 750},
        "cap_DUL": {"so_ong": 5, "so_tao_cap": 12, "ong_PVC_D": 49},
    }

    dt = tinh_du_toan_so_bo(sample_res, "Đồng bằng Nam Bộ", "Cấp III",
                             beam_params=sample_beam)

    th = dt["tong_hop"]
    ph = dt["phan_tich"]
    print("=" * 60)
    print("DỰ TOÁN SƠ BỘ — Super-T 38.2m, 12m, 1 nhịp")
    print("=" * 60)
    print(f"  HM1 KCN:   {th['hm1_KCN']/1e6:8.3f} tỷ VND")
    print(f"  HM2 MTM:   {th['hm2_MTM']/1e6:8.3f} tỷ VND")
    print(f"  HM3 Móng:  {th['hm3_MON']/1e6:8.3f} tỷ VND")
    print(f"  HM4 BMC:   {th['hm4_BMC']/1e6:8.3f} tỷ VND")
    print(f"  HM5 Phụ:   {th['hm5_PHU']/1e6:8.3f} tỷ VND")
    print(f"  ─────────────────────────────")
    print(f"  Tổng XL:   {th['tong_xay_lap']/1e6:8.3f} tỷ VND")
    print(f"  +Phụ phí:  {(th['tong_truoc_VAT']-th['tong_xay_lap'])/1e6:8.3f} tỷ VND")
    print(f"  +VAT 10%:  {th['VAT']/1e6:8.3f} tỷ VND")
    print(f"  ─────────────────────────────")
    print(f"  TỔNG ĐT:   {th['tong_dau_tu']/1e6:8.3f} tỷ VND")
    print(f"  Mặt cầu:   {ph['A_cau_m2']:.0f} m²")
    print(f"  Suất ĐT:   {ph['suat_dau_tu_trieu_m2']:.2f} triệu VND/m²")
    print(f"  {ph['canh_bao']}")
