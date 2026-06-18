"""
Module 00 — Địa chất Loader
Đọc file Excel địa chất (template hoặc file thực tế tương thích) và trả về
dữ liệu chuẩn hóa để phục vụ Module 08 phân loại móng cọc.

Định dạng file hỗ trợ:
  • Template_DiaChat.xlsx (Toado_HK 7 cột, HKx 8 cột mới)
  • HK-VVD.xlsx (Toado_HK 4 cột, SPT dùng "-" cho ô trống)

Các cột mới trong sheet HKx (v2):
  LOAI_DAT — phân loại đất (Sét, Cát, Đá tươi, …)
  IL       — chỉ số chảy sét (0–2.0), chỉ cho đất dính
  RQD      — chỉ số chất lượng đá (%), chỉ cho đá
  GHI_CHU_LOP — ghi chú đặc điểm lớp
"""

import re
import os
import numpy as np
import openpyxl

# ═══════════════════════════════════════════════════════════════════════════════
# HẰNG SỐ
# ═══════════════════════════════════════════════════════════════════════════════
_SPT_WEAK_THRESHOLD   = 10   # N < 10 → đất yếu
_SPT_STOP_THRESHOLD   = 15   # N ≥ 15 liên tiếp 2 mẫu → kết thúc vùng yếu
_SPT_GOOD_THRESHOLD   = 20   # N ≥ 20 liên tiếp 3 mẫu → lớp chịu lực
_SPT_STOP_CONSECUTIVE = 2
_SPT_GOOD_CONSECUTIVE = 3

# Ngưỡng SPT-N cho lớp tựa mũi theo loại đất
_SPT_NGUONG_DINH = 30   # đất dính: Sét, Sét pha
_SPT_NGUONG_ROI  = 40   # đất rời: Cát, Cát pha, Cát lẫn sỏi, Sỏi cuội
_RQD_NGUONG      = 60   # đá: RQD > 60%
_CHIEU_DAY_MIN   = 4.0  # chiều dày tối thiểu lớp tựa mũi (m)

# Nhóm loại đất
_DAT_DINH = {"Sét", "Sét pha"}
_DAT_ROI  = {"Cát", "Cát pha", "Cát lẫn sỏi", "Sỏi cuội"}
_DAT_DA   = {"Đá phong hóa", "Đá tươi"}


# ═══════════════════════════════════════════════════════════════════════════════
# HÀM PHỤ TRỢ
# ═══════════════════════════════════════════════════════════════════════════════
def _parse_do_sau(s):
    """
    '2.0 - 2.45' → (2.0, 2.45)
    '2.0'        → (2.0, 2.45)   [fallback: +0.45 m]
    """
    if s is None:
        return None, None
    s = str(s).strip()
    m = re.match(r"^([\d.]+)\s*[-–]\s*([\d.]+)\s*$", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    try:
        v = float(s)
        return v, v + 0.45
    except (ValueError, TypeError):
        return None, None


def _parse_spt_value(raw):
    """Chuyển đổi giá trị ô thành int N hoặc None. '-' / trống → None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("-", "", "nan", "None", "-"):
        return None
    try:
        f = float(s)
        if f < 0 or f > 100:
            return None
        return int(round(f))
    except (ValueError, TypeError):
        return None


def _normalize_hk_name(name):
    """Strip whitespace + chuẩn hóa để so sánh tên hố khoan."""
    return str(name).strip() if name else ""


def _parse_optional_float(val, lo=None, hi=None):
    """Chuyển giá trị tùy chọn thành float, kiểm tra phạm vi."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("-", "", "nan", "None"):
        return None
    try:
        f = float(s)
        if lo is not None and f < lo:
            return None
        if hi is not None and f > hi:
            return None
        return f
    except (ValueError, TypeError):
        return None


def _infer_loai_dat_from_name(ten_lop: str) -> str:
    """Suy ra LOAI_DAT từ TEN_LOP nếu cột LOAI_DAT không có."""
    t = str(ten_lop).lower().strip()
    if any(k in t for k in ["đất đắp", "san lấp", "đắp", "lớp k"]):
        return "Đất đắp"
    if any(k in t for k in ["than bùn", "hữu cơ", "bùn than"]):
        return "Đất hữu cơ"
    if "đá tươi" in t:
        return "Đá tươi"
    if "đá phong hóa" in t or ("đá" in t and "phong hóa" in t):
        return "Đá phong hóa"
    if "đá" in t:
        return "Đá phong hóa"
    if "sỏi cuội" in t or ("sỏi" in t and "cuội" in t):
        return "Sỏi cuội"
    if "cát lẫn sỏi" in t or ("cát" in t and "sỏi" in t):
        return "Cát lẫn sỏi"
    if "sét pha" in t or "bùn sét pha" in t:
        return "Sét pha"
    if "sét" in t or "bùn" in t:
        return "Sét"
    if "cát pha" in t:
        return "Cát pha"
    if "cát" in t:
        return "Cát"
    return ""


# ═══════════════════════════════════════════════════════════════════════════════
# 1. TÍNH ĐẶC TRƯNG TỔNG HỢP (SPT-N)
# ═══════════════════════════════════════════════════════════════════════════════
def _compute_characteristics(spt_list):
    """
    Tính các đặc trưng địa kỹ thuật từ chuỗi SPT-N.

    Parameters
    ----------
    spt_list : list of (start_depth, end_depth, N|None)
               Đã sắp xếp theo start_depth tăng dần.

    Returns
    -------
    dict với các trường:
        chieu_sau_lop_yeu       : chiều sâu kết thúc lớp yếu (m từ mặt đất) hoặc None
        spt_n_lop_yeu           : SPT-N TB vùng yếu hoặc None
        chieu_sau_lop_tot       : chiều sâu bắt đầu lớp chịu lực (m) hoặc None
        spt_n_lop_tot           : SPT-N TB lớp chịu lực hoặc None
        do_sau_ket_thuc_khoan   : độ sâu cuối cùng có N hợp lệ (m) hoặc None
        dia_chat_du_lieu_co     : True nếu đủ dữ liệu để phân loại móng tin cậy
    """
    valid = [(s, e, n) for s, e, n in spt_list if n is not None]

    base = {
        "chieu_sau_lop_yeu":     None,
        "spt_n_lop_yeu":         None,
        "chieu_sau_lop_tot":     None,
        "spt_n_lop_tot":         None,
        "do_sau_ket_thuc_khoan": None,
        "dia_chat_du_lieu_co":   False,
    }
    if not valid:
        return base

    base["do_sau_ket_thuc_khoan"] = round(max(e for _, e, _ in valid), 2)
    base["dia_chat_du_lieu_co"]   = len(valid) >= 5

    # ── Vùng yếu: SPT < 10, dừng khi 2 mẫu liên tiếp N ≥ 15 ─────────────────
    weak_started    = False
    weak_n_values   = []
    last_weak_end   = None
    consec_strong   = 0

    for s, e, n in valid:
        if n < _SPT_WEAK_THRESHOLD:
            weak_started  = True
            weak_n_values.append(n)
            last_weak_end = e
            consec_strong = 0
        elif weak_started:
            if n >= _SPT_STOP_THRESHOLD:
                consec_strong += 1
                if consec_strong >= _SPT_STOP_CONSECUTIVE:
                    break
            else:
                consec_strong = 0

    if weak_started and weak_n_values:
        base["chieu_sau_lop_yeu"] = round(last_weak_end, 2)
        base["spt_n_lop_yeu"]     = round(float(np.mean(weak_n_values)), 1)

    # ── Lớp chịu lực: 3 mẫu liên tiếp N ≥ 20 ─────────────────────────────────
    consec_good     = 0
    candidate_start = None
    good_start      = None

    for s, e, n in valid:
        if n >= _SPT_GOOD_THRESHOLD:
            if consec_good == 0:
                candidate_start = s
            consec_good += 1
            if consec_good >= _SPT_GOOD_CONSECUTIVE:
                good_start = candidate_start
                break
        else:
            consec_good     = 0
            candidate_start = None

    if good_start is not None:
        base["chieu_sau_lop_tot"] = round(good_start, 2)
        good_n = [n for s, e, n in valid if s >= good_start]
        if good_n:
            base["spt_n_lop_tot"] = round(float(np.mean(good_n)), 1)

    return base


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TÌM LỚP TỰA MŨI VÀ ĐẶC TRƯNG MỚI
# ═══════════════════════════════════════════════════════════════════════════════
def _avg_spt_for_layer(lop, spt_list_full, z_mat_dat):
    """
    Tính SPT-N trung bình cho một lớp đất dựa trên cao độ lớp và dữ liệu SPT.

    lop          : dict với cao_do_dinh, cao_do_day (cao độ tuyệt đối, m)
    spt_list_full: list of (start_depth, end_depth, N|None) — độ sâu từ mặt đất
    z_mat_dat    : cao độ mặt đất (m)
    """
    if not spt_list_full or z_mat_dat is None:
        return None

    # Chuyển cao độ lớp sang độ sâu từ mặt đất
    depth_dinh = z_mat_dat - lop["cao_do_dinh"]
    depth_day  = z_mat_dat - lop["cao_do_day"]

    n_values = []
    for s, e, n in spt_list_full:
        if n is None:
            continue
        mid = (s + e) / 2
        if depth_dinh - 0.5 <= mid <= depth_day + 0.5:
            n_values.append(n)

    if not n_values:
        return None
    return round(float(np.mean(n_values)), 1)


def _find_bearing_layer(lop_dat, spt_list_full, z_mat_dat):
    """
    Tìm lớp tựa mũi cọc dự kiến và các cờ kỹ thuật liên quan.

    Parameters
    ----------
    lop_dat       : list[dict] — danh sách lớp đất đã có loai_dat, IL, RQD
    spt_list_full : list of (start_depth, end_depth, N|None)
    z_mat_dat     : cao độ mặt đất (m) để quy đổi sang độ sâu

    Returns
    -------
    dict:
        lop_tua_mui_de_xuat  : dict mô tả lớp tựa mũi, hoặc None
        co_lop_tua_mui_du_kien: bool
        co_da_phong_hoa      : bool
        cao_do_da_phong_hoa  : float|None
        co_da_tuoi           : bool
        cao_do_da_tuoi       : float|None
        co_set_chay          : bool — có sét IL > 0.6 bất kỳ đâu trong hố
        co_da_moi_co_giua    : bool — đá phong hóa/mồ côi không phải lớp cuối
    """
    result = {
        "lop_tua_mui_de_xuat":   None,
        "co_lop_tua_mui_du_kien": False,
        "co_da_phong_hoa":        False,
        "cao_do_da_phong_hoa":    None,
        "co_da_tuoi":             False,
        "cao_do_da_tuoi":         None,
        "co_set_chay":            False,
        "co_da_moi_co_giua":      False,
    }
    if not lop_dat:
        return result

    n_layers = len(lop_dat)

    # ── Quét toàn bộ lớp để đặt cờ ──────────────────────────────────────────
    for i, lop in enumerate(lop_dat):
        loai = lop.get("loai_dat", "") or _infer_loai_dat_from_name(lop.get("ten_lop", ""))
        IL   = lop.get("IL")
        RQD  = lop.get("RQD")
        ghi  = str(lop.get("ghi_chu_lop", "") or "").lower()

        # Sét chảy
        if loai in _DAT_DINH and IL is not None and IL > 0.6:
            result["co_set_chay"] = True

        # Đá phong hóa (ghi nhận lần đầu)
        if loai == "Đá phong hóa" and not result["co_da_phong_hoa"]:
            result["co_da_phong_hoa"]      = True
            result["cao_do_da_phong_hoa"]  = lop["cao_do_dinh"]
            # Nếu có lớp đất phi-đá phía dưới → đá ở giữa (mồ côi/xen kẹp)
            if i < n_layers - 1:
                has_soil_below = any(
                    lop_dat[j].get("loai_dat", "") not in _DAT_DA
                    for j in range(i + 1, n_layers)
                )
                if has_soil_below:
                    result["co_da_moi_co_giua"] = True

        # Đá mồ côi từ GHI_CHU_LOP
        if "mồ côi" in ghi or "mo coi" in ghi:
            result["co_da_moi_co_giua"] = True

        # Đá tươi (ghi nhận lần đầu)
        if loai == "Đá tươi" and not result["co_da_tuoi"]:
            result["co_da_tuoi"]     = True
            result["cao_do_da_tuoi"] = lop["cao_do_dinh"]

    # ── Tìm lớp tựa mũi ──────────────────────────────────────────────────────
    for lop in lop_dat:
        loai     = lop.get("loai_dat", "") or _infer_loai_dat_from_name(lop.get("ten_lop", ""))
        IL       = lop.get("IL")
        RQD      = lop.get("RQD")
        chieu_day = lop.get("chieu_day", 0) or 0

        # Điều kiện 3: chiều dày ≥ 4m
        if chieu_day < _CHIEU_DAY_MIN:
            continue

        # Điều kiện 2: loại trừ sét chảy
        if loai in _DAT_DINH and IL is not None and IL > 0.6:
            continue

        # Tính SPT-N trung bình của lớp
        z = z_mat_dat if z_mat_dat is not None else 0.0
        spt_avg = _avg_spt_for_layer(lop, spt_list_full, z)

        # Điều kiện 1: ngưỡng SPT/RQD theo loại đất
        qualifies = False
        if loai in _DAT_DINH:
            qualifies = spt_avg is not None and spt_avg > _SPT_NGUONG_DINH
        elif loai in _DAT_ROI:
            qualifies = spt_avg is not None and spt_avg > _SPT_NGUONG_ROI
        elif loai == "Đá phong hóa":
            qualifies = RQD is not None and RQD > _RQD_NGUONG
        elif loai == "Đá tươi":
            # Đá tươi luôn đủ điều kiện tựa mũi CKN
            qualifies = True

        if qualifies:
            do_sau = round(z - lop["cao_do_dinh"], 2) if z is not None else None
            result["lop_tua_mui_de_xuat"] = {
                "ten_lop":    lop["ten_lop"],
                "loai_dat":   loai,
                "IL":         IL,
                "RQD":        RQD,
                "cao_do_dinh": lop["cao_do_dinh"],
                "cao_do_day":  lop["cao_do_day"],
                "chieu_day":   chieu_day,
                "do_sau_dinh": do_sau,
                "spt_n_tb":    spt_avg,
            }
            result["co_lop_tua_mui_du_kien"] = True
            break

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# 3. HÀM CHÍNH: ĐỌC FILE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def load_geology_excel(file_path):
    """
    Đọc file Excel địa chất (template hoặc tương thích).

    Parameters
    ----------
    file_path : str | Path — đường dẫn file .xlsx / .xlsm

    Returns
    -------
    dict:
        ho_khoan_list    : list[dict]  — dữ liệu từng hố khoan
        validation_errors: list[dict]  — {ten_sheet, hang, cot, mo_ta_loi}
        dac_trung_tong_hop: dict       — {ten_hk: characteristics_dict}
    """
    result = {
        "ho_khoan_list":      [],
        "validation_errors":  [],
        "dac_trung_tong_hop": {},
    }
    errors = result["validation_errors"]

    # ── Mở workbook ────────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        errors.append({"ten_sheet": "-", "hang": "-", "cot": "-",
                        "mo_ta_loi": f"Không mở được file: {exc}"})
        return result

    sheets = wb.sheetnames

    # ── Sheet Toado_HK ─────────────────────────────────────────────────────────
    if "Toado_HK" not in sheets:
        errors.append({"ten_sheet": "Toado_HK", "hang": "-", "cot": "-",
                        "mo_ta_loi": "Thiếu sheet bắt buộc: Toado_HK"})
        return result

    ws_tk  = wb["Toado_HK"]
    header = [
        str(c.value).strip().upper() if c.value else ""
        for c in next(ws_tk.iter_rows(min_row=1, max_row=1))
    ]
    col_map = {h: i for i, h in enumerate(header)}  # tên → index 0-based

    hk_list_raw = []
    for row in ws_tk.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        ten = _normalize_hk_name(row[col_map.get("HO_KHOAN", 0)])
        if not ten:
            continue

        def _safe_float(key, default=None):
            idx = col_map.get(key)
            if idx is None:
                return default
            try:
                return float(row[idx])
            except (TypeError, ValueError):
                return default

        def _safe_str(key, default=""):
            idx = col_map.get(key)
            if idx is None:
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        hk_list_raw.append({
            "ten":         ten,
            "X":           _safe_float("X"),
            "Y":           _safe_float("Y"),
            "Z":           _safe_float("Z"),
            "Z_mat_nuoc":  _safe_float("Z_MAT_NUOC"),
            "ly_trinh":    _safe_float("LY_TRINH"),
            "ghi_chu":     _safe_str("GHI_CHU"),
        })

    hk_names = {hk["ten"] for hk in hk_list_raw}

    # ── Sheet SPT ──────────────────────────────────────────────────────────────
    spt_raw_by_hk = {}

    if "SPT" not in sheets:
        errors.append({"ten_sheet": "SPT", "hang": "-", "cot": "-",
                        "mo_ta_loi": "Thiếu sheet bắt buộc: SPT"})
    else:
        ws_spt  = wb["SPT"]
        spt_hdr = [
            str(c.value).strip() if c.value else ""
            for c in next(ws_spt.iter_rows(min_row=1, max_row=1))
        ]

        spt_col_map = {}
        for i, h in enumerate(spt_hdr):
            if i == 0:
                continue
            m = re.match(r"^(.+?)\s*\(N\)\s*$", h, re.IGNORECASE)
            key = m.group(1).strip() if m else h.strip()
            if key:
                spt_col_map[key] = i

        for name in hk_names:
            if name not in spt_col_map:
                errors.append({
                    "ten_sheet": "SPT",
                    "hang": "header",
                    "cot": f"{name} (N)",
                    "mo_ta_loi": f"Hố khoan '{name}' thiếu cột SPT tương ứng",
                })
        for key in spt_col_map:
            if key not in hk_names:
                errors.append({
                    "ten_sheet": "SPT",
                    "hang": "header",
                    "cot": f"{key} (N)",
                    "mo_ta_loi": f"Cột SPT '{key}' không tương ứng hố khoan nào trong Toado_HK",
                })

        for name in spt_col_map:
            spt_raw_by_hk[name] = []

        for row_no, row in enumerate(ws_spt.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            start_d, end_d = _parse_do_sau(row[0])
            if start_d is None:
                continue

            for name, ci in spt_col_map.items():
                raw = row[ci] if ci < len(row) else None
                n   = _parse_spt_value(raw)

                if raw is not None and str(raw).strip() not in ("-", ""):
                    try:
                        f = float(str(raw).strip())
                        if not (0 <= f <= 100):
                            errors.append({
                                "ten_sheet": "SPT",
                                "hang": str(row_no),
                                "cot": f"{name} (N)",
                                "mo_ta_loi": f"Giá trị SPT-N = {raw} ngoài khoảng 0–100",
                            })
                    except (ValueError, TypeError):
                        pass

                spt_raw_by_hk[name].append((start_d, end_d, n))

    # ── Sheets phân lớp địa chất ────────────────────────────────────────────────
    SYSTEM_SHEETS = {"HUONG_DAN", "Toado_HK", "SPT"}
    soil_by_hk    = {}

    for sname in sheets:
        if sname in SYSTEM_SHEETS:
            continue
        ws_soil = wb[sname]
        raw_hdr = [
            str(c.value).strip().upper() if c.value else ""
            for c in next(ws_soil.iter_rows(min_row=1, max_row=1))
        ]
        soil_col = {h: i for i, h in enumerate(raw_hdr)}

        if sname not in hk_names:
            errors.append({
                "ten_sheet": sname,
                "hang": "-", "cot": "-",
                "mo_ta_loi": f"Sheet '{sname}' không có trong danh sách Toado_HK",
            })

        layers = []
        for row_no, row in enumerate(ws_soil.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v for v in row if v is not None):
                continue

            def _get(key, default=None):
                idx = soil_col.get(key)
                return row[idx] if (idx is not None and idx < len(row)) else default

            ten_raw = _get("TEN_LOP")
            if ten_raw is None:
                continue
            ten_lop = str(ten_raw).strip()
            if not ten_lop:
                continue

            try:
                cao_dinh = float(_get("CAO_DO_DINH_LOP", 0) or 0)
                cao_day  = float(_get("CAO_DO_DAY_LOP",  0) or 0)
            except (TypeError, ValueError):
                cao_dinh = cao_day = 0.0

            # Đọc các cột mới (tương thích ngược nếu không có)
            loai_dat_raw = _get("LOAI_DAT")
            loai_dat = str(loai_dat_raw).strip() if loai_dat_raw else ""
            if not loai_dat:
                loai_dat = _infer_loai_dat_from_name(ten_lop)

            il_val  = _parse_optional_float(_get("IL"),  lo=0.0, hi=2.0)
            rqd_val = _parse_optional_float(_get("RQD"), lo=0.0, hi=100.0)
            ghi_lop = str(_get("GHI_CHU_LOP", "") or "").strip()

            layers.append({
                "ten_lop":     ten_lop,
                "cao_do_dinh": cao_dinh,
                "cao_do_day":  cao_day,
                "chieu_day":   round(cao_dinh - cao_day, 2),
                "mo_ta":       str(_get("MO_TA", "") or "").strip(),
                "loai_dat":    loai_dat,
                "IL":          il_val,
                "RQD":         rqd_val,
                "ghi_chu_lop": ghi_lop,
            })

        soil_by_hk[sname] = layers

    # ── Ghép dữ liệu + validation ──────────────────────────────────────────────
    for hk in hk_list_raw:
        ten      = hk["ten"]
        lop_dat  = soil_by_hk.get(ten, [])
        spt_list = spt_raw_by_hk.get(ten, [])

        # Validation: đỉnh lớp đầu ≈ Z mặt đất
        if lop_dat and hk.get("Z") is not None:
            diff = abs(lop_dat[0]["cao_do_dinh"] - hk["Z"])
            if diff > 0.10:
                errors.append({
                    "ten_sheet": ten,
                    "hang": "2",
                    "cot": "CAO_DO_DINH_LOP",
                    "mo_ta_loi": (
                        f"Cao độ đỉnh lớp đầu ({lop_dat[0]['cao_do_dinh']:.2f} m) "
                        f"lệch {diff:.2f} m so với Z mặt đất ({hk['Z']:.2f} m)"
                    ),
                })

        # Validation: đáy lớp trên = đỉnh lớp dưới
        for i in range(len(lop_dat) - 1):
            gap = abs(lop_dat[i]["cao_do_day"] - lop_dat[i + 1]["cao_do_dinh"])
            if gap > 0.05:
                errors.append({
                    "ten_sheet": ten,
                    "hang": str(i + 3),
                    "cot": "CAO_DO_DINH_LOP",
                    "mo_ta_loi": (
                        f"Lớp {i+1} ({lop_dat[i]['ten_lop']}) đáy = {lop_dat[i]['cao_do_day']:.2f} m "
                        f"≠ đỉnh lớp {i+2} ({lop_dat[i+1]['ten_lop']}) = {lop_dat[i+1]['cao_do_dinh']:.2f} m "
                        f"(lệch {gap:.2f} m)"
                    ),
                })

        hk["lop_dat"]  = lop_dat
        hk["spt_data"] = [(s, n) for s, e, n in spt_list if n is not None]
        result["ho_khoan_list"].append(hk)

        # Đặc trưng SPT cổ điển
        chars = _compute_characteristics(spt_list)

        # Đặc trưng tựa mũi và cờ kỹ thuật mới
        z_mat = hk.get("Z")
        bearing_info = _find_bearing_layer(lop_dat, spt_list, z_mat)
        chars.update(bearing_info)

        # Gắn thêm Z vào chars để Module 08 dùng
        chars["Z"] = z_mat

        result["dac_trung_tong_hop"][ten] = chars

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# 4. KIỂM CHỨNG DỮ LIỆU
# ═══════════════════════════════════════════════════════════════════════════════
def validate_geology_data(data):
    """
    Chạy lại các kiểm tra trên dữ liệu đã load và trả về báo cáo text.

    Parameters
    ----------
    data : dict — kết quả từ load_geology_excel()

    Returns
    -------
    str — báo cáo kiểm chứng (multiline text)
    """
    errs    = data.get("validation_errors", [])
    hk_list = data.get("ho_khoan_list", [])

    lines = ["=" * 60, "BÁO CÁO KIỂM CHỨNG DỮ LIỆU ĐỊA CHẤT", "=" * 60]
    lines.append(f"Tổng số hố khoan đọc được : {len(hk_list)}")

    if not errs:
        lines.append("Kết quả kiểm chứng       : ✅ ĐẠT — Không phát hiện lỗi")
    else:
        lines.append(f"Số cảnh báo / lỗi        : {len(errs)}")
        lines.append("")
        lines.append("CHI TIẾT:")
        for e in errs:
            lines.append(
                f"  [{e['ten_sheet']}] Hàng {e['hang']} | Cột {e['cot']}"
            )
            lines.append(f"    → {e['mo_ta_loi']}")

    lines.append("")
    lines.append("TÓM TẮT HỐ KHOAN:")
    dt = data.get("dac_trung_tong_hop", {})
    for hk in hk_list:
        ten = hk["ten"]
        d   = dt.get(ten, {})
        def _fmt(val, unit="m"):
            return f"{val}{unit}" if val is not None else "—"

        z_str   = f"{hk['Z']:.2f}m" if hk.get("Z") is not None else "—"
        yeu_str = f"{_fmt(d.get('chieu_sau_lop_yeu'))}(N={_fmt(d.get('spt_n_lop_yeu','—'), '')})"
        tot_str = f"{_fmt(d.get('chieu_sau_lop_tot'))}(N={_fmt(d.get('spt_n_lop_tot','—'), '')})"
        end_str = _fmt(d.get("do_sau_ket_thuc_khoan"))

        # Thông tin lớp tựa mũi mới
        lop_tua = d.get("lop_tua_mui_de_xuat")
        tua_str = "Chưa tìm được"
        if lop_tua:
            tua_str = (
                f"{lop_tua['ten_lop']} ({lop_tua['loai_dat']}) "
                f"tại -{lop_tua['do_sau_dinh']:.1f}m, "
                f"dày {lop_tua['chieu_day']:.1f}m"
            )

        flags = []
        if d.get("co_set_chay"):      flags.append("⚠️ Sét chảy")
        if d.get("co_da_moi_co_giua"): flags.append("⚠️ Đá mồ côi")
        if d.get("co_da_phong_hoa"):   flags.append("🪨 Đá PH")
        if d.get("co_da_tuoi"):        flags.append("🪨 Đá tươi")

        lines.append(f"  {ten}: Z={z_str}  Lớp_yếu={yeu_str}  Lớp_tốt={tot_str}  Khoan_đến={end_str}")
        lines.append(f"       Lớp_tựa_mũi: {tua_str}")
        if flags:
            lines.append(f"       Cờ: {' | '.join(flags)}")

    lines.append("=" * 60)
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# SELF-TEST
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys, pathlib
    sys.stdout.reconfigure(encoding="utf-8")

    _DIR = pathlib.Path(__file__).parent
    test_file = _DIR / "Survey" / "HCM-MB" / "HK-VVD.xlsx"

    print(f"Kiểm chứng với: {test_file}")
    d = load_geology_excel(str(test_file))

    print(f"\nSố hố khoan : {len(d['ho_khoan_list'])}")
    for hk in d["ho_khoan_list"]:
        print(f"  [{hk['ten']}]  X={hk['X']}  Y={hk['Y']}  Z={hk['Z']}")
        for lop in hk["lop_dat"]:
            print(
                f"    Lớp {lop['ten_lop']:8s}  {lop['cao_do_dinh']:7.2f} → {lop['cao_do_day']:7.2f}"
                f"  ({lop['chieu_day']:.2f}m)  [{lop['loai_dat']}]"
                f"  IL={lop['IL']}  RQD={lop['RQD']}"
            )

    print("\nĐặc trưng tổng hợp:")
    for ten, c in d["dac_trung_tong_hop"].items():
        print(f"  {ten}:")
        lop_tua = c.get("lop_tua_mui_de_xuat")
        if lop_tua:
            print(f"    Lớp tựa mũi: {lop_tua['ten_lop']} ({lop_tua['loai_dat']}) "
                  f"tại -{lop_tua['do_sau_dinh']:.1f}m")
        print(f"    co_set_chay={c.get('co_set_chay')}  co_da_moi={c.get('co_da_moi_co_giua')}")

    print()
    print(validate_geology_data(d))
