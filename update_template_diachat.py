"""
Script cập nhật Template_DiaChat.xlsx: bổ sung 4 cột mới vào sheets HK
LOAI_DAT, IL, RQD, GHI_CHU_LOP
Chạy một lần: python update_template_diachat.py
"""
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

_DIR = os.path.dirname(os.path.abspath(__file__))
_TPL = os.path.join(_DIR, "Data", "Template_DiaChat.xlsx")

LOAI_DAT_VALUES = [
    "Sét", "Sét pha", "Cát", "Cát pha",
    "Cát lẫn sỏi", "Sỏi cuội",
    "Đá phong hóa", "Đá tươi",
    "Đất đắp", "Đất hữu cơ",
]

NEW_COLS = [
    ("LOAI_DAT",    "Phân loại đất (bắt buộc)"),
    ("IL",          "Chỉ số chảy sét (0–2.0)"),
    ("RQD",         "Chỉ số chất lượng đá (%)"),
    ("GHI_CHU_LOP", "Ghi chú đặc điểm lớp"),
]

# Màu header cột mới
_NEW_HDR_FILL  = PatternFill("solid", fgColor="FFF2CC")  # vàng nhạt
_OLD_HDR_FILL  = PatternFill("solid", fgColor="4472C4")  # xanh dương
_FONT_HDR_NEW  = Font(bold=True, color="7F4F00", size=10)
_FONT_HDR_OLD  = Font(bold=True, color="FFFFFF", size=10)
_ALIGN_C       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN          = Side(border_style="thin", color="BFBFBF")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _infer_loai_dat(ten_lop: str) -> str:
    """Suy ra LOAI_DAT từ tên lớp."""
    t = str(ten_lop).lower().strip()
    if any(k in t for k in ["đất đắp", "lớp k", "san lấp", "đắp"]):
        return "Đất đắp"
    if any(k in t for k in ["than bùn", "hữu cơ", "bùn than"]):
        return "Đất hữu cơ"
    if "đá tươi" in t or ("đá" in t and "phong hóa" not in t and "tươi" in t):
        return "Đá tươi"
    if "đá phong hóa" in t or ("đá" in t and "phong hóa" in t):
        return "Đá phong hóa"
    if "đá" in t:
        return "Đá phong hóa"
    if "sỏi cuội" in t or ("sỏi" in t and "cuội" in t):
        return "Sỏi cuội"
    if "cát lẫn sỏi" in t or ("cát" in t and "sỏi" in t):
        return "Cát lẫn sỏi"
    if "sét pha" in t or "bùn sét pha" in t:
        return "Sét pha"
    if "sét" in t or "đất sét" in t or "bùn" in t:
        return "Sét"
    if "cát pha" in t:
        return "Cát pha"
    if "cát" in t:
        return "Cát"
    return ""


def _format_header_cell(cell, is_new: bool):
    cell.fill      = _NEW_HDR_FILL if is_new else _OLD_HDR_FILL
    cell.font      = _FONT_HDR_NEW if is_new else _FONT_HDR_OLD
    cell.alignment = _ALIGN_C
    cell.border    = _BORDER


def update_hk_sheet(ws, wb):
    """Bổ sung 4 cột mới vào sheet phân lớp HKx."""
    # Đọc header dòng 1
    header_row = [c.value for c in ws[1]]
    header_upper = [str(h).strip().upper() if h else "" for h in header_row]

    # Xác định cột hiện có (dùng chữ hoa để so sánh)
    existing_cols = {h: i + 1 for i, h in enumerate(header_upper)}

    # Tìm cột bắt đầu chèn thêm (sau MO_TA hoặc sau cột cuối cùng có nội dung)
    insert_after = existing_cols.get("MO_TA", max(existing_cols.values(), default=4))

    # Tên cột mới cần thêm (bỏ qua nếu đã tồn tại)
    to_add = [(c, d) for c, d in NEW_COLS if c not in existing_cols]
    if not to_add:
        print(f"  [{ws.title}] Tất cả cột mới đã tồn tại — bỏ qua.")
        return

    # Chèn cột sau MO_TA
    start_col = insert_after + 1
    for offset, (col_name, _) in enumerate(to_add):
        col_idx = start_col + offset
        # Chèn cột trống
        ws.insert_cols(col_idx)
        # Ghi header
        hdr_cell = ws.cell(row=1, column=col_idx, value=col_name)
        _format_header_cell(hdr_cell, is_new=True)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 16 if col_name == "LOAI_DAT" else 10

    # Đọc lại header sau khi chèn
    header_row2 = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]
    col_map2 = {h: i + 1 for i, h in enumerate(header_row2)}

    # Điền dữ liệu mẫu cho các dòng có TEN_LOP
    ten_col = col_map2.get("TEN_LOP", 1)
    ld_col  = col_map2.get("LOAI_DAT")
    il_col  = col_map2.get("IL")
    rqd_col = col_map2.get("RQD")

    for row_idx in range(2, ws.max_row + 1):
        ten_cell = ws.cell(row=row_idx, column=ten_col)
        ten_val  = str(ten_cell.value).strip() if ten_cell.value else ""
        if not ten_val:
            continue

        # LOAI_DAT: suy ra từ TEN_LOP nếu ô đang trống
        if ld_col:
            ld_cell = ws.cell(row=row_idx, column=ld_col)
            if not ld_cell.value:
                ld_cell.value = _infer_loai_dat(ten_val)
            ld_cell.alignment = _ALIGN_C

        # IL: sample values nếu trống
        if il_col:
            il_cell = ws.cell(row=row_idx, column=il_col)
            if il_cell.value is None and ld_col:
                ld_val = str(ws.cell(row=row_idx, column=ld_col).value or "")
                if ld_val in ("Sét", "Sét pha"):
                    il_cell.value = 0.3  # giá trị ví dụ
            il_cell.alignment = _ALIGN_C

        # RQD: sample values nếu trống
        if rqd_col:
            rqd_cell = ws.cell(row=row_idx, column=rqd_col)
            if rqd_cell.value is None and ld_col:
                ld_val = str(ws.cell(row=row_idx, column=ld_col).value or "")
                if ld_val == "Đá phong hóa":
                    rqd_cell.value = 45
                elif ld_val == "Đá tươi":
                    rqd_cell.value = 82
            rqd_cell.alignment = _ALIGN_C

    # Data validation cho LOAI_DAT
    if ld_col and ws.max_row > 1:
        col_letter = openpyxl.utils.get_column_letter(ld_col)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(LOAI_DAT_VALUES) + '"',
            allow_blank=True,
            showErrorMessage=True,
            error="Chọn từ danh sách cho phép",
            errorTitle="Giá trị không hợp lệ",
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{col_letter}2:{col_letter}{max(ws.max_row, 50)}"

    print(f"  [{ws.title}] Đã thêm {len(to_add)} cột: {[c for c,_ in to_add]}")


def update_huong_dan(ws):
    """Bổ sung mô tả các cột mới vào cuối sheet HUONG_DAN."""
    # Tìm dòng cuối có nội dung
    last_row = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                last_row = cell.row

    start = last_row + 2

    lines = [
        ("", ""),
        ("═══════════════ CÁC CỘT MỚI TRONG SHEET HKx ═══════════════", ""),
        ("", ""),
        ("Cột LOAI_DAT — Phân loại đất (BẮT BUỘC)", ""),
        ("", "Ý nghĩa: Phân loại đất theo nhóm để hệ thống áp dụng ngưỡng SPT-N phù hợp."),
        ("", "Giá trị cho phép: Sét | Sét pha | Cát | Cát pha | Cát lẫn sỏi | Sỏi cuội | Đá phong hóa | Đá tươi | Đất đắp | Đất hữu cơ"),
        ("", "Ví dụ: Lớp 'Bùn sét dẻo nhão' → Sét pha; Lớp 'Cát vừa chặt vừa' → Cát"),
        ("", ""),
        ("Cột IL — Chỉ số chảy (Liquidity Index)", ""),
        ("", "Ý nghĩa: Chỉ áp dụng cho lớp đất dính (Sét, Sét pha). Để trống cho loại đất khác."),
        ("", "Công thức: IL = (w - wP) / Ip, trong đó w = độ ẩm tự nhiên, wP = giới hạn dẻo, Ip = chỉ số dẻo"),
        ("", "Phạm vi: 0.0 (cứng) → 1.0 (dẻo chảy) → 2.0 (chảy); IL > 0.6 = sét mềm/dẻo chảy"),
        ("", "Ví dụ điển hình: Sét cứng IL≈0.1; Sét dẻo cứng IL≈0.35; Sét dẻo mềm IL≈0.55; Sét chảy IL≈0.85"),
        ("", ""),
        ("Cột RQD — Rock Quality Designation (%)", ""),
        ("", "Ý nghĩa: Chỉ áp dụng cho Đá phong hóa và Đá tươi. Để trống cho đất."),
        ("", "Công thức: RQD = Tổng chiều dài lõi khoan nguyên vẹn ≥ 10cm / Tổng chiều dài khoan × 100%"),
        ("", "Phạm vi: 0–25% = Đá rất kém; 25–50% = Đá kém; 50–75% = Đá trung bình; 75–90% = Đá tốt; >90% = Đá rất tốt"),
        ("", "Ngưỡng cọc khoan nhồi tựa mũi vào đá: RQD ≥ 60% (TCVN 10304:2014)"),
        ("", "Ví dụ: Đá phong hóa nặng RQD≈30%; Đá phong hóa nhẹ RQD≈65%; Đá tươi RQD≈85%"),
        ("", ""),
        ("Cột GHI_CHU_LOP — Ghi chú đặc điểm lớp (tùy chọn)", ""),
        ("", "Mục đích: Ghi nhận đặc điểm kỹ thuật đặc biệt ảnh hưởng đến thi công cọc."),
        ("", "Ví dụ quan trọng: 'Có đá mồ côi lẫn' → hệ thống cảnh báo cọc ép không hạ qua được"),
        ("", "Các từ khóa hệ thống nhận dạng: đá mồ côi; hang động; vật liệu hữu cơ; bùn lỏng"),
        ("", ""),
        ("LƯU Ý QUAN TRỌNG:", ""),
        ("", "1. LOAI_DAT là cột BẮT BUỘC. Không điền sẽ ảnh hưởng đến độ chính xác tư vấn móng cọc."),
        ("", "2. IL > 0.6 → Module 08 sẽ cảnh báo và có thể loại trừ cọc ép tựa mũi vào lớp đó."),
        ("", "3. RQD > 60% → lớp đá được xem là phù hợp để cọc khoan nhồi ngàm vào (TCVN 10304)."),
        ("", "4. 'Đá mồ côi' trong GHI_CHU_LOP → Module 08 bắt buộc dùng cọc khoan nhồi."),
    ]

    title_fill = PatternFill("solid", fgColor="4472C4")
    sec_fill   = PatternFill("solid", fgColor="D6E4F0")
    title_font = Font(bold=True, color="FFFFFF", size=11)
    sec_font   = Font(bold=True, color="1F3864", size=10)

    for i, (col_a, col_b) in enumerate(lines):
        r = start + i
        if col_a:
            c = ws.cell(row=r, column=1, value=col_a)
            if "═" in col_a:
                c.fill = title_fill
                c.font = title_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            elif col_a.startswith("Cột ") or col_a.startswith("LƯU Ý"):
                c.fill = sec_fill
                c.font = sec_font
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        if col_b:
            cell_b = ws.cell(row=r, column=2, value=col_b)
            cell_b.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 80
    print(f"  [HUONG_DAN] Đã bổ sung {len(lines)} dòng hướng dẫn cột mới.")


def add_sample_data_hk(ws, hk_idx):
    """Bổ sung dữ liệu mẫu cho các cột mới (nếu sheet có ít dữ liệu)."""
    # Sample data cho 3 HK khác nhau
    samples = {
        1: [
            # ten_lop (pattern), LOAI_DAT, IL, RQD, ghi_chu
            ("Đất đắp", "Đất đắp", None, None, "Lớp san lấp mặt bằng"),
            ("Sét", "Sét", 1.15, None, "Sét dẻo chảy, IL > 0.6 — chú ý cọc ép"),
            ("Cát pha", "Cát pha", None, None, ""),
            ("Cát chịu lực", "Cát", None, None, ""),
            ("Sét cứng", "Sét", 0.08, None, "Lớp tựa mũi dự kiến"),
        ],
        2: [
            ("Sét pha", "Sét pha", 0.42, None, ""),
            ("Cát vừa", "Cát", None, None, ""),
            ("Cát lẫn sỏi", "Cát lẫn sỏi", None, None, "Có đá mồ côi lẫn"),
            ("Sét cứng", "Sét", 0.12, None, "Lớp chịu lực tốt"),
        ],
        3: [
            ("Sét pha", "Sét pha", 0.50, None, ""),
            ("Cát chặt", "Cát", None, None, ""),
            ("Đá phong hóa", "Đá phong hóa", None, 42, "Đá phong hóa nặng"),
            ("Đá tươi", "Đá tươi", None, 82, "Đá granite tươi — tựa mũi CKN"),
        ],
    }
    data_for_hk = samples.get(hk_idx, [])
    if not data_for_hk:
        return

    # Đọc lại header
    header_row = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]
    col_map    = {h: i + 1 for i, h in enumerate(header_row)}
    ten_col    = col_map.get("TEN_LOP", 1)
    ld_col     = col_map.get("LOAI_DAT")
    il_col     = col_map.get("IL")
    rqd_col    = col_map.get("RQD")
    ghi_col    = col_map.get("GHI_CHU_LOP")

    for row_idx in range(2, ws.max_row + 1):
        ten_val = str(ws.cell(row=row_idx, column=ten_col).value or "").strip()
        if not ten_val:
            continue
        # Match với sample data theo thứ tự
        row_offset = row_idx - 2
        if row_offset < len(data_for_hk):
            _, ld, il, rqd, ghi = data_for_hk[row_offset]
            if ld_col  and not ws.cell(row=row_idx, column=ld_col).value:
                ws.cell(row=row_idx, column=ld_col).value = ld
            if il_col  and il is not None and not ws.cell(row=row_idx, column=il_col).value:
                ws.cell(row=row_idx, column=il_col).value = il
            if rqd_col and rqd is not None and not ws.cell(row=row_idx, column=rqd_col).value:
                ws.cell(row=row_idx, column=rqd_col).value = rqd
            if ghi_col and ghi and not ws.cell(row=row_idx, column=ghi_col).value:
                ws.cell(row=row_idx, column=ghi_col).value = ghi


def main():
    if not os.path.exists(_TPL):
        print(f"❌ Không tìm thấy file: {_TPL}")
        sys.exit(1)

    print(f"Mở template: {_TPL}")
    wb = openpyxl.load_workbook(_TPL)
    sheets = wb.sheetnames
    print(f"Sheets hiện có: {sheets}")

    # Cập nhật HUONG_DAN
    if "HUONG_DAN" in sheets:
        update_huong_dan(wb["HUONG_DAN"])
    else:
        print("  [HUONG_DAN] Sheet không tồn tại — bỏ qua")

    # Cập nhật các sheet HK
    SYSTEM_SHEETS = {"HUONG_DAN", "Toado_HK", "SPT"}
    hk_idx = 1
    for sname in sheets:
        if sname in SYSTEM_SHEETS:
            continue
        update_hk_sheet(wb[sname], wb)
        add_sample_data_hk(wb[sname], hk_idx)
        hk_idx += 1

    wb.save(_TPL)
    print(f"\n✅ Đã lưu: {_TPL}")
    print("   Template đã được cập nhật với 4 cột mới: LOAI_DAT, IL, RQD, GHI_CHU_LOP")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
