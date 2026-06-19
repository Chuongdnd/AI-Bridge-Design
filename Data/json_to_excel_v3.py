"""
CÔNG CỤ CHUYỂN JSON → EXCEL TRAIN DATASET v3
============================================
Cách dùng:
  1. Paste JSON từ NotebookLM vào biến JSON_DATA (hỗ trợ cả ```json ... ```)
  2. Chọn SHEET_TARGET: "01" | "02" | "03" | "04" | "05" | "06" | "07" | "08"
  3. Chạy: python json_to_excel_v3.py
  4. Nếu trùng tên cầu → UPDATE tại chỗ; cầu mới → APPEND từ dòng trống tiếp theo.

Yêu cầu: pip install openpyxl

JSON key chuẩn theo từng sheet (khớp template trong Prompt_0X_*.txt):
  01: ten_cau · ly_trinh · tinh_thanh_pho · nam_xay_dung · chu_dau_tu
      don_vi_tu_van · tieu_chuan_tk · tong_muc_dau_tu · trang_thai
      nguon_du_lieu · confidence · ngay_trich_xuat
  02: ten_cau · vung_dia_ly · loai_vuot · loai_duong · cap_ky_thuat · vtk
      i_doc · ban_kinh_dc · tai_trong_tk · loai_cau_kc · so_do_nhip
      tong_so_nhip · l_cau · goc_xien · b_cau · don_nguyen · bc
      so_lan_xe_ct · rong_1_lan
  03: ten_cau · q1pct · v_dong_chay · b_song · h_nuoc_tb · he_so_thu_hep
      xoi_chung · xoi_cuc_bo · cao_do_day_xoi · loai_day_song
      che_do_thuy_trieu · cap_song · b_tinh_khong · h_tinh_khong
      cao_do_mncn · cao_do_mntt · cao_do_mntc · cao_do_mntn
  04: ten_cau · so_lop_dc · ten_lop_dat_yeu · day_lop_yeu · spt_n
      ten_lop_dat_tot · sau_lop_chiu_luc · spt_n_tot
  05: ten_cau · loai_mong · so_coc · chieu_dai_coc · duong_kinh_coc
      cao_do_mat_be · kt_be_coc · cap_bt_coc · pp_tc_coc · sct_coc
  06: ten_cau · loai_tru · h_tru · rong_tru · dai_tru · so_cot
      hinh_dang_tru · xa_mu_tru · cap_bt_tru · so_tru_giua · kc_tim_tru
      loai_mo · h_mo · dai_than_mo · tuong_canh_mo · tuong_chan_dat
      duong_dan · vi_tri · bao_ve_xoi
  07: ten_cau · loai_dam · vat_lieu_dam · pp_thi_cong · tinh_lien_tuc
      cap_bt_dam · l_tt · l_dam · h_dam · b_dam · ti_le_lh · so_dam
      kc_tim_dam · hang_dau_dam · so_dam_ngang
  08: ten_cau · day_bmc · day_lop_phu · loai_goi · loai_khe
      rong_xe_chay · rong_le_bh · dai_pc_giua · cau_tao_mc
      loai_lan_can · lo_thoat_nuoc · chieu_sang
"""

import json, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ════════════════════════════════════════════════════════════
# ① PASTE JSON VÀO ĐÂY
# ════════════════════════════════════════════════════════════
JSON_DATA = '''
{
  "ten_cau": "Cầu Kênh Ngang",
  "loai_mong": "Cọc đóng BTCT",
  "so_coc": 30,
  "chieu_dai_coc": 60.0,
  "duong_kinh_coc": 600,
  "cao_do_mat_be": 1.0,
  "kt_be_coc": "12.0m x 6.0m x 1.5m",
  "cap_bt_coc": "C60",
  "pp_tc_coc": "Đóng búa",
  "sct_coc": null
}
'''

# ② CHỌN SHEET
SHEET_TARGET = "05"   # 05=Móng | 06=Mố-Trụ | 07=Dầm | 08=Mặt cầu

# ③ FILE EXCEL
EXCEL_FILE = "Bridge_Train_Dataset_v3.xlsx"

# ════════════════════════════════════════════════════════════
# MAP: tên cột Excel (row 3) → JSON key(s)
# Key chính lấy từ row 7 trong Excel; alias để tương thích các
# JSON cũ hoặc tên do NotebookLM sinh ra.
# ════════════════════════════════════════════════════════════
COL_MAP = {
    # ── SHEET 01: Thông tin dự án ─────────────────────────────
    "Tên cầu":                              "ten_cau",
    "Lý trình":                             "ly_trinh",
    "Tỉnh / Thành phố":                     "tinh_thanh_pho",
    "Năm xây dựng / hoàn thành":            ["nam_xay_dung", "nam_xd"],
    "Chủ đầu tư":                           "chu_dau_tu",
    "Đơn vị tư vấn TK":                     "don_vi_tu_van",
    "Tiêu chuẩn thiết kế":                  ["tieu_chuan_tk", "tieu_chuan_thiet_ke"],
    "Tổng mức đầu tư":                      "tong_muc_dau_tu",
    "Trạng thái hiện tại":                  "trang_thai",
    "Nguồn dữ liệu":                        ["nguon_du_lieu", "__LM__"],
    "Mức độ tin cậy":                       "confidence",
    "Ngày trích xuất":                      ["ngay_trich_xuat", "__DATE__"],
    # ── SHEET 02: Tổng thể ────────────────────────────────────
    "Vùng địa lý":                          "vung_dia_ly",
    "Loại đối tượng vượt":                  ["loai_vuot", "loai_doi_tuong_vuot"],
    "Loại đường trên cầu":                  ["loai_duong", "loai_duong_tren_cau"],
    "Cấp kỹ thuật đường":                   ["cap_ky_thuat", "cap_ky_thuat_duong"],
    "Vận tốc thiết kế Vtk":                 ["vtk", "van_toc_tk"],
    "Độ dốc dọc cầu i_doc":                 ["i_doc", "doc_doc_i"],
    "Bán kính đường cong đứng R":           ["ban_kinh_dc", "ban_kinh_duong_cong"],
    "Tải trọng thiết kế":                   ["tai_trong_tk", "tai_trong_thiet_ke"],
    "Loại cầu theo kết cấu":                ["loai_cau_kc", "loai_cau_ket_cau"],
    "Sơ đồ nhịp":                           "so_do_nhip",
    "Tổng số nhịp":                         "tong_so_nhip",
    "Chiều dài toàn cầu L_cầu":             ["l_cau", "chieu_dai_cau"],
    "Góc xiên":                             "goc_xien",
    "Bề rộng toàn cầu B_cầu":              ["b_cau", "be_rong_cau"],
    "Số đơn nguyên cầu":                    ["don_nguyen", "so_don_nguyen"],
    "Tổng bề rộng cầu Bc":                  ["bc", "bc_kho_cau"],
    "Số làn xe (chi tiết)":                 ["so_lan_xe_ct", "so_lan_xe"],
    "Bề rộng 1 làn xe":                     ["rong_1_lan", "rong_lan_xe"],
    # ── SHEET 03: Thủy văn ────────────────────────────────────
    "Lưu lượng thiết kế Q1%":               ["q1pct", "q_1_pct"],
    "Vận tốc dòng chảy TB":                 "v_dong_chay",
    "Chiều rộng lòng sông B_sông":          ["b_song", "chieu_rong_song"],
    "Chiều sâu nước TB h_tb":               "h_nuoc_tb",
    "Hệ số thu hẹp dòng chảy":             "he_so_thu_hep",
    "Chiều sâu xói chung":                  "xoi_chung",
    "Chiều sâu xói cục bộ trụ/mố":         "xoi_cuc_bo",
    "Cao độ đáy xói thiết kế":              ["cao_do_day_xoi", "cao_do_day_xoi_tk"],
    "Loại lòng sông / đáy kênh":            ["loai_day_song", "loai_long_song"],
    "Chế độ thủy triều":                    "che_do_thuy_trieu",
    "Cấp sông ĐTNĐ":                        ["cap_song", "cap_song_dtnd"],
    "Bề rộng tĩnh không B":                 "b_tinh_khong",
    "Chiều cao tĩnh không H":               "h_tinh_khong",
    "Cao độ MNCN (H1%)":                    ["cao_do_mncn", "cao_do_mncn_h1_pct"],
    "Cao độ MNTT (H5%)":                    ["cao_do_mntt", "cao_do_mntt_h5_pct"],
    "Cao độ MNTC (H10%)":                   "cao_do_mntc",
    "Cao độ MNTN (H98%)":                   "cao_do_mntn",
    # ── SHEET 04: Địa chất ────────────────────────────────────
    "Số lớp địa chất":                      ["so_lop_dc", "so_lop_dia_chat"],
    "Tên lớp đất yếu":                      ["ten_lop_dat_yeu", "lop_dat_chu_dao"],
    "Chiều dày lớp đất yếu":               ["day_lop_yeu", "day_lop_dat_yeu"],
    "SPT-N trung bình lớp đất yếu":        ["spt_n", "spt-n_tb"],
    "Tên lớp đất tốt":                      "ten_lop_dat_tot",
    "Chiều sâu lớp chịu lực tốt":          "sau_lop_chiu_luc",
    "SPT-N trung bình lớp đất tốt":        "spt_n_tot",
    # ── SHEET 05: Móng ────────────────────────────────────────
    "⚑ LOẠI MÓNG":                          "loai_mong",
    "Số lượng cọc / bệ":                    ["so_coc", "so_coc_be"],
    "Chiều dài cọc":                        ["chieu_dai_coc", "chieu_dai_cọc", "chieu_dai_coc_m"],
    "Đường kính cọc":                       ["duong_kinh_coc", "duong_kin_coc"],
    "Cao độ mặt bệ cọc":                    ["cao_do_mat_be", "cao_do_mat_be_coc"],
    "Kích thước bệ cọc":                    ["kt_be_coc", "kich_thuoc_be"],
    "Cấp bê tông cọc":                      "cap_bt_coc",
    "Phương pháp thi công cọc":             ["pp_tc_coc", "pp_thi_cong_coc"],
    "Sức chịu tải thiết kế / cọc":         ["sct_coc", "suc_chiu_tai_coc"],
    # ── SHEET 06: Mố – Trụ ────────────────────────────────────
    "⚑ LOẠI TRỤ":                           "loai_tru",
    "Chiều cao thân trụ H_trụ":             ["h_tru", "chieu_cao_tru"],
    "Chiều rộng thân trụ (dọc cầu)":       ["rong_tru", "rong_tru_doc_cau"],
    "Chiều dài thân trụ (ngang cầu)":      ["dai_tru", "dai_tru_ngang_cau"],
    "Số cột / trụ (khung)":                ["so_cot", "so_cot_tru"],
    "Hình dạng mặt cắt thân trụ":          ["hinh_dang_tru", "hinh_dang_than_tru"],
    "Kích thước xà mũ trụ":                ["xa_mu_tru", "kich_thuoc_xa_mu"],
    "Cấp bê tông trụ":                     "cap_bt_tru",
    "Số trụ giữa / cầu":                   "so_tru_giua",
    "Khoảng cách tim trụ":                 ["kc_tim_tru", "khoang_cach_tim_tru"],
    "Loại mố":                             "loai_mo",
    "Chiều cao mố H_mố":                   ["h_mo", "chieu_cao_mo"],
    "Chiều dài thân mố (ngang cầu)":       ["dai_than_mo", "dai_mo_ngang"],
    "Kích thước tường cánh mố":            ["tuong_canh_mo", "kich_thuoc_tuong_canh"],
    "Loại tường chắn đất sau mố":          ["tuong_chan_dat", "loai_tuong_chan_dat"],
    "Chiều dài đường dẫn vào cầu":         ["duong_dan", "chieu_dai_duong_dan"],
    "Vị trí mố/trụ trong sông":            ["vi_tri", "vi_tri_tru_mo"],
    "Bảo vệ chống xói":                    ["bao_ve_xoi", "bao_ve_chong_xoi"],
    # ── SHEET 07: Kết cấu nhịp ────────────────────────────────
    "⚑ LOẠI DẦM":                           "loai_dam",
    "Vật liệu dầm":                         "vat_lieu_dam",
    "Phương pháp thi công":                 ["pp_thi_cong", "pp_thi_cong_nhip"],
    "Tính liên tục":                        "tinh_lien_tuc",
    "Cấp bê tông dầm":                      ["cap_bt_dam", "cap_be_tong_dam"],
    "Chiều dài nhịp tính toán L_tt":        ["l_tt", "l_tt_m", "chieu_dai_nhip_tt"],
    "Chiều dài dầm L_dầm":                  ["l_dam", "l_dam_m", "chieu_dai_dam"],
    "Chiều cao dầm H_dầm":                  ["h_dam", "h_dam_m", "chieu_cao_dam"],
    "Bề rộng dầm B_dầm":                    "b_dam",
    "Tỉ lệ L/H":                            ["ti_le_lh", "ti_le_l_h"],
    "Số lượng dầm / mặt cắt":              ["so_dam", "so_dam_mat_cat", "so_luong_dam"],
    "Khoảng cách tim dầm":                  ["kc_tim_dam", "khoang_cach_tim_dam"],
    "Phần hẫng đầu dầm (overhang)":        ["hang_dau_dam", "hang_dau"],
    "Số dầm ngang / nhịp":                  "so_dam_ngang",
    "Chiều dày bản mặt cầu":               ["day_bmc", "day_ban_mat_cau"],
    "Loại gối cầu":                         ["loai_goi", "loai_goi_cau"],
    "Loại khe co giãn":                     ["loai_khe", "loai_khe_co_gian"],
    "Chiều dày lớp phủ mặt cầu":           ["day_lop_phu", "day_lop_phu_mat_cau"],
    # ── SHEET 08: Mặt cầu ─────────────────────────────────────
    # (Chiều dày bản mặt cầu, Loại gối cầu, Loại khe co giãn
    #  trùng tên với sheet 07 → cùng mapping, Python lấy cái sau)
    "Chiều rộng phần xe chạy":              ["rong_xe_chay", "rong_phan_xe_chay"],
    "Chiều rộng lề bộ hành":               ["rong_le_bh", "rong_le_bo_hanh"],
    "Dải phân cách giữa":                  ["dai_pc_giua", "dai_phan_cach_giua"],
    "Cấu tạo lớp mặt cầu":                ["cau_tao_mc", "cau_tao_mat_cau"],
    "Loại lan can / hộ lan":               ["loai_lan_can", "loai_ho_lan"],
    "Hệ thống thoát nước mặt cầu":        ["lo_thoat_nuoc", "he_thong_thoat_nuoc"],
    "Chiếu sáng cầu":                      ["chieu_sang", "chieu_sang_cau"],
}

SHEET_NAMES = {
    "01": "01_Thông_tin_DA",
    "02": "02_Tổng thể",
    "03": "03_Thủy văn",
    "04": "04_Địa chất",
    "05": "05_Móng",
    "06": "06_Mố – Trụ",
    "07": "07_Kết cấu nhịp",
    "08": "08_Mặt cầu",
}

# ── Styles ────────────────────────────────────────────────────────────────────
def fill(h): return PatternFill('solid', fgColor=h, start_color=h)
thin = Side(style='thin', color='BBBBBB')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
CONF = {'high':'C6EFCE', 'medium':'FFF2CC', 'low':'FFD6D6'}

def get_val(rec, col_name):
    from datetime import date
    key = COL_MAP.get(col_name)
    if key is None: return None
    if isinstance(key, list):
        for k in key:
            if k == "__LM__":   return "NotebookLM"
            if k == "__DATE__": return date.today().strftime("%d/%m/%Y")
            v = rec.get(k)
            if v is not None: return v
        return None
    if key == "__LM__":   return "NotebookLM"
    if key == "__DATE__": return date.today().strftime("%d/%m/%Y")
    return rec.get(key)

def main():
    try:
        cleaned = JSON_DATA.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1]
        if cleaned.endswith("```"):
            cleaned = cleaned.rsplit("```", 1)[0]
        records = json.loads(cleaned)
        if isinstance(records, dict): records = [records]
    except json.JSONDecodeError as e:
        print(f"❌ JSON lỗi: {e}"); sys.exit(1)

    sheet_name = SHEET_NAMES.get(SHEET_TARGET)
    if not sheet_name:
        print(f"❌ SHEET_TARGET phải là 01/02/03/04/05/06/07/08"); sys.exit(1)

    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"❌ Không tìm thấy: {EXCEL_FILE}"); sys.exit(1)

    ws = wb[sheet_name]

    # Header luôn ở row 3 trong tất cả các sheet
    col_positions = {}
    for cell in ws[3]:
        if cell.value and str(cell.value).strip() not in ('', 'None'):
            col_positions[str(cell.value).strip()] = cell.column

    if not col_positions:
        print("❌ Không tìm được header ở row 3"); sys.exit(1)

    # Scan dữ liệu hiện có: build map tên cầu → dòng, tìm dòng trống tiếp theo
    START = 9
    ten_cau_col = col_positions.get("Tên cầu")
    existing_rows = {}  # ten_cau → row_number

    next_row = START
    for r in range(START, START+500):
        sample = [ws.cell(row=r, column=c).value
                  for c in list(col_positions.values())[:6]]
        if all(v in (None, '') for v in sample):
            next_row = r
            break
        if ten_cau_col:
            val = ws.cell(row=r, column=ten_cau_col).value
            if val:
                existing_rows[str(val).strip()] = r

    print(f"✅ Sheet     : {sheet_name}")
    print(f"✅ Cầu hiện có: {len(existing_rows)}")
    print(f"✅ Append từ : dòng {next_row}")
    print(f"✅ Số cầu mới: {len(records)}")
    print()

    written = updated = new_offset = 0
    for ri, rec in enumerate(records):
        ten_cau = str(rec.get('ten_cau', '')).strip()
        if ten_cau and ten_cau in existing_rows:
            row = existing_rows[ten_cau]
            action = "UPDATE"
        else:
            row = next_row + new_offset
            new_offset += 1
            action = "NEW   "

        rb   = 'FAFAFA' if ri%2==0 else 'F0F4FA'
        conf = str(rec.get('confidence','medium')).lower()

        for col_name, col_idx in col_positions.items():
            val  = get_val(rec, col_name)
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font      = Font(name='Arial', size=10)

            if '⚑' in col_name:
                cell.fill = fill('FFE4E1')
                cell.font = Font(name='Arial', size=10, bold=True, color='8B0000')
            elif col_name in ('Tên cầu','Lý trình','Tỉnh / Thành phố'):
                cell.fill = fill(CONF.get(conf,'FAFAFA'))
            else:
                cell.fill = fill(rb)

        ws.row_dimensions[row].height = 18
        written += 1
        if action.startswith("UPDATE"):
            updated += 1
        name   = rec.get('ten_cau', f'#{ri+1}')
        filled = sum(1 for k,v in rec.items() if v is not None)
        null_  = sum(1 for k in col_positions if get_val(rec,k) is None)
        print(f"  [{ri+1:02d}] [{action}] dòng {row:<4} {name:<35} filled={filled:2d}  null_cols={null_:2d}")

    wb.save(EXCEL_FILE)
    print(f"\n✅ Đã xử lý {written} cầu → '{sheet_name}'  (mới: {written-updated}, cập nhật: {updated})  |  File: {EXCEL_FILE}")

if __name__ == '__main__':
    main()
