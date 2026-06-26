"""
Dashboard so sánh ba phương án kết cấu nhịp — Bước 6B
"""
import copy
import io
import json
import math
import importlib.util
import os
from datetime import datetime

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="So sánh Phương án",
    page_icon="📊",
    layout="wide",
)

_DIR  = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_DIR)

# ══════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════
PA_KEYS   = ["PA1", "PA2", "PA3"]
PA_COLORS = {"PA1": "#1f77b4", "PA2": "#9467bd", "PA3": "#2ca02c"}
PA_BG     = {"PA1": "#dbeafe", "PA2": "#ede9fe", "PA3": "#dcfce7"}
PA_LABELS = {
    "PA1": "PA1 — Tối ưu chi phí",
    "PA2": "PA2 — Tối ưu mỹ quan",
    "PA3": "PA3 — AI khuyến nghị",
}
SCORE_LABELS = {
    "A1": "L/H hợp lý",    "A2": "Nhịp phù hợp", "A3": "Thi công",
    "A4": "Vật liệu",      "A5": "Độ bền",
    "B1": "Chi phí tổng",  "B2": "Suất đầu tư",   "B3": "Hiệu quả KT",
    "C1": "Tỉ lệ thẩm mỹ","C2": "Kiểu dáng",
}
_BEAM_SPAN_OK = {
    "Super-T":      (25, 50),
    "Dầm I":        (15, 40),
    "T ngược":      (12, 25),
    "Dầm T":        (15, 35),
    "Dầm bản rỗng": (8,  25),
}
_BEAM_AES = {
    "Super-T": 9, "Dầm T": 7, "T ngược": 6, "Dầm I": 5, "Dầm bản rỗng": 4,
}

# ══════════════════════════════════════════════════════════════════
# MODULE LOADING
# ══════════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner=False)
def _load_du_toan_mod():
    spec = importlib.util.spec_from_file_location(
        "du_toan", os.path.join(_ROOT, "12-DuToan_SoBo.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

# ══════════════════════════════════════════════════════════════════
# CROSS-SECTION DRAWING
# ══════════════════════════════════════════════════════════════════

def _g(mc: dict, vn_key: str, en_key: str, default: float) -> float:
    return float(mc.get(vn_key) or mc.get(en_key) or default)


def _supert_pts(mc: dict, H: float):
    bf = _g(mc, "B_canh_tren",   "bf_top",  2440)
    tf = _g(mc, "H_canh_tren",   "tf_top",  200)
    bw = _g(mc, "B_bung_top",    "bw",      200)
    bd = _g(mc, "B_canh_duoi",   "bf_bot",  700)
    td = _g(mc, "H_canh_duoi",   "tf_bot",  200)
    cx = bf / 2
    xs = [0, bf, bf, cx+bw/2, cx+bw/2, cx+bd/2,
          cx+bd/2, cx-bd/2, cx-bd/2, cx-bw/2, cx-bw/2, 0, 0]
    ys = [H, H, H-tf, H-tf, td, td,
          0, 0, td, td, H-tf, H-tf, H]
    return xs, ys


def _damI_pts(mc: dict, H: float):
    bf_t = _g(mc, "B_canh_tren",  "bf_top", 600)
    tf_t = _g(mc, "H_canh_tren",  "tf_top", 130)
    bf_b = _g(mc, "B_canh_duoi",  "bf_bot", 500)
    tf_b = _g(mc, "H_canh_duoi",  "tf_bot", 130)
    bw   = float(mc.get("B_bung") or mc.get("bw") or 200)
    cx   = max(bf_t, bf_b) / 2
    xs = [cx-bf_t/2, cx+bf_t/2, cx+bf_t/2, cx+bw/2, cx+bw/2, cx+bf_b/2,
          cx+bf_b/2, cx-bf_b/2, cx-bf_b/2, cx-bw/2, cx-bw/2, cx-bf_t/2, cx-bf_t/2]
    ys = [H, H, H-tf_t, H-tf_t, tf_b, tf_b,
          0, 0, tf_b, tf_b, H-tf_t, H-tf_t, H]
    return xs, ys


def _tnguoc_pts(mc: dict, H: float):
    bf_b = _g(mc, "B_canh_duoi", "bf_bot", 1200)
    tf_b = _g(mc, "H_canh_duoi", "tf_bot", 200)
    bw   = float(mc.get("B_bung") or mc.get("bw") or 200)
    cx   = bf_b / 2
    xs = [0, bf_b, bf_b, cx+bw/2, cx+bw/2, cx-bw/2, cx-bw/2, 0, 0]
    ys = [0, 0, tf_b, tf_b, H, H, tf_b, tf_b, 0]
    return xs, ys


def _dam_rect_pts(B: float, H: float):
    return [0, B, B, 0, 0], [0, 0, H, H, 0]


def _section_fig(pa: dict, color: str = "#1f77b4") -> go.Figure:
    bp    = pa.get("beam_params", {})
    mc    = bp.get("MC_AA", {})
    H     = float(bp.get("H", 1750))
    loai  = pa.get("loai_dam", "Super-T")
    B     = float(bp.get("B", 1800))

    try:
        if loai == "Super-T":
            xs, ys = _supert_pts(mc, H)
        elif loai in ("Dầm I",):
            xs, ys = _damI_pts(mc, H)
        elif loai in ("T ngược",):
            xs, ys = _tnguoc_pts(mc, H)
        else:
            xs, ys = _dam_rect_pts(B, H)
    except Exception:
        xs, ys = _dam_rect_pts(B, H)

    B_max = max(xs) if xs else B
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=xs, y=ys,
        fill="toself", fillcolor=color + "33",
        line=dict(color=color, width=2),
        mode="lines", showlegend=False,
        hoverinfo="skip",
    ))
    fig.update_layout(
        height=200,
        margin=dict(l=4, r=4, t=4, b=4),
        xaxis=dict(range=[-B_max * 0.08, B_max * 1.08],
                   visible=False, scaleanchor="y", scaleratio=1),
        yaxis=dict(range=[-H * 0.08, H * 1.12], visible=False),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig

# ══════════════════════════════════════════════════════════════════
# SCORING (100 points)
# ══════════════════════════════════════════════════════════════════

def _score_pa(pa: dict, all_costs: list) -> dict:
    bp    = pa.get("beam_params", {})
    loai  = pa.get("loai_dam", "Super-T")
    L     = float(pa.get("L_nhip", 38.2))
    H_m   = float(bp.get("H", 1750)) / 1000.0
    cost  = float(pa.get("tong_dau_tu", 0))
    suat  = float(pa.get("suat_dau_tu_trieu", 11.0))

    # ── A: Kỹ thuật (60) ──
    LH = L / max(H_m, 0.1)
    if 18 <= LH <= 25:              a1 = 12
    elif 15 <= LH < 18 or 25 < LH <= 30: a1 = 9
    elif 12 <= LH < 15 or 30 < LH <= 35: a1 = 6
    else:                            a1 = 3

    lo, hi = _BEAM_SPAN_OK.get(loai, (15, 45))
    if lo <= L <= hi:               a2 = 12
    elif lo * 0.8 <= L <= hi * 1.1: a2 = 8
    else:                            a2 = 4

    a3 = 11 if loai in ("Super-T", "Dầm I") else (9 if loai == "Dầm T" else 7)

    n_dam = int(pa.get("n_dam", 6))
    bc    = float(pa.get("bc", 12.0))
    A_sec = float(pa.get("A_tiet_dien_m2", 0.808))
    ratio = A_sec * n_dam / max(bc, 1)
    if ratio < 0.45:   a4 = 12
    elif ratio < 0.60: a4 = 9
    elif ratio < 0.75: a4 = 6
    else:              a4 = 3

    cng = str(pa.get("cong_nghe", "DUL"))
    a5  = 11 if "DUL" in cng else 7

    # ── B: Kinh tế (30) ──
    min_c = min(all_costs) if all_costs else cost
    max_c = max(all_costs) if all_costs else cost
    span  = max(max_c - min_c, 1)
    b1 = max(0, min(15, round(15 * (1 - (cost - min_c) / span))))
    b2 = (10 if 8 <= suat <= 12 else
          7  if 7 <= suat <= 14 else
          4  if suat <= 16 else 1)
    b3 = max(0, min(5, round(5 * (1 - (cost - min_c) / span))))

    # ── C: Mỹ quan (10) ──
    aes = _BEAM_AES.get(loai, 5)
    c1  = round(5 * aes / 10)
    c2  = max(0, min(5, round(5 * (1 - abs(LH - 20) / 20))))

    total = a1 + a2 + a3 + a4 + a5 + b1 + b2 + b3 + c1 + c2
    return {
        "A1": a1, "A2": a2, "A3": a3, "A4": a4, "A5": a5,
        "B1": b1, "B2": b2, "B3": b3,
        "C1": c1, "C2": c2,
        "diem_ky_thuat": a1 + a2 + a3 + a4 + a5,
        "diem_kinh_te":  b1 + b2 + b3,
        "diem_my_quan":  c1 + c2,
        "tong_diem":     total,
    }

# ══════════════════════════════════════════════════════════════════
# BUILD DEMO / REAL PA LIST
# ══════════════════════════════════════════════════════════════════

def _make_pa_entry(spec: dict, dt_mod, base_res: dict) -> dict:
    res = copy.deepcopy(base_res)
    res.update(spec.get("res_override", {}))
    bp = spec["beam_params"]
    try:
        dt = dt_mod.tinh_du_toan_so_bo(res, vung_mien="Đồng bằng Nam Bộ",
                                        cap_duong="Cấp III", beam_params=bp)
        th = dt["tong_hop"]
        pa_dt = dt["phan_tich"]
    except Exception:
        th    = {k: 0 for k in ["hm1_KCN","hm2_MTM","hm3_MON","hm4_BMC",
                                 "hm5_PHU","tong_xay_lap","tong_dau_tu"]}
        pa_dt = {"A_cau_m2": 0, "suat_dau_tu_trieu_m2": 0}

    H_m = float(bp.get("H", 1750)) / 1000.0
    L   = spec["L_nhip"]
    return {
        "key":       spec["key"],
        "loai_dam":  spec["loai_dam"],
        "L_nhip":    L,
        "n_nhip":    spec["n_nhip"],
        "n_dam":     spec["n_dam"],
        "bc":        spec.get("bc", 12.0),
        "cong_nghe": bp.get("cong_nghe", ""),
        "H_dam_m":   round(H_m, 3),
        "B_dam_m":   round(float(bp.get("B", 1800)) / 1000.0, 3),
        "S_dam_m":   round(float(bp.get("S", 2200)) / 1000.0, 3),
        "LH_ratio":  round(L / max(H_m, 0.1), 1),
        "beam_params": bp,
        "A_tiet_dien_m2": spec.get("A_sec", 0.808),
        "th":             th,
        "pa_dt":          pa_dt,
        "tong_dau_tu":    float(th.get("tong_dau_tu", 0)),
        "suat_dau_tu_trieu": float(pa_dt.get("suat_dau_tu_trieu_m2", 0)),
        "A_cau_m2":       float(pa_dt.get("A_cau_m2", 0)),
    }


def _demo_pa_list(dt_mod) -> list:
    base_res = {
        "bc": 12.0,
        "geo_logic": {"L_cau": 38.2},
        "kcn_result": {
            "so_luong_dam": 6, "loai_dam": "Super-T",
            "chieu_dai": 38.2, "chieu_cao_dam": 1.75,
            "khoang_cach_dam": 2.2, "tong_so_nhip": 1,
        },
        "tru_result": {"so_tru": 0, "chieu_cao_tru": 5.0},
        "mong_result": {"D_coc": 0.8, "L_coc": 20.0, "so_coc_tren_be": 5,
                        "loai_mong": "Cọc khoan nhồi"},
        "lop_phu_result": {"chieu_day_ao_duong": 0.12},
    }
    specs = [
        {
            "key": "PA1", "loai_dam": "Super-T",
            "L_nhip": 38.2, "n_nhip": 1, "n_dam": 6, "bc": 12.0,
            "A_sec": 0.808,
            "beam_params": {
                "loai_dam": "Super-T", "L": 38200, "H": 1750, "S": 2200, "B": 2440,
                "cong_nghe": "DUL căng sau",
                "MC_AA": {"bf_top": 2440, "tf_top": 200, "bw": 200, "hw": 1250,
                          "bf_bot": 700, "tf_bot": 200},
                "cap_DUL": {"so_ong": 5, "D_ong": 100, "n_tao": 12, "luc_cang_kN": 185},
            },
            "res_override": {},
        },
        {
            "key": "PA2", "loai_dam": "T ngược",
            "L_nhip": 18.0, "n_nhip": 2, "n_dam": 6, "bc": 12.0,
            "A_sec": 0.440,
            "beam_params": {
                "loai_dam": "T ngược", "L": 18000, "H": 1100, "S": 2000, "B": 1200,
                "cong_nghe": "BTCT thường",
                "MC_AA": {"bf_bot": 1200, "tf_bot": 200, "bw": 200, "hw": 700},
                "cap_DUL": {"so_ong": 0},
            },
            "res_override": {
                "geo_logic": {"L_cau": 36.0},
                "kcn_result": {
                    "so_luong_dam": 6, "loai_dam": "T ngược",
                    "chieu_dai": 18.0, "chieu_cao_dam": 1.1,
                    "khoang_cach_dam": 2.0, "tong_so_nhip": 2,
                },
            },
        },
        {
            "key": "PA3", "loai_dam": "Dầm I",
            "L_nhip": 33.0, "n_nhip": 1, "n_dam": 6, "bc": 12.0,
            "A_sec": 0.620,
            "beam_params": {
                "loai_dam": "Dầm I", "L": 33000, "H": 1650, "S": 2200, "B": 1800,
                "cong_nghe": "DUL căng trước",
                "MC_AA": {"bf_top": 600, "tf_top": 130, "bw": 200, "hw": 1250,
                          "bf_bot": 500, "tf_bot": 130},
                "cap_DUL": {"so_ong": 4, "D_ong": 90, "n_tao": 10, "luc_cang_kN": 155},
            },
            "res_override": {
                "geo_logic": {"L_cau": 33.0},
                "kcn_result": {
                    "so_luong_dam": 6, "loai_dam": "Dầm I",
                    "chieu_dai": 33.0, "chieu_cao_dam": 1.65,
                    "khoang_cach_dam": 2.2, "tong_so_nhip": 1,
                },
            },
        },
    ]
    return [_make_pa_entry(s, dt_mod, base_res) for s in specs]

# ══════════════════════════════════════════════════════════════════
# SENSITIVITY RECALCULATION
# ══════════════════════════════════════════════════════════════════

def _sensitivity_total(th: dict, k_bt: float, k_thep: float,
                       k_nc: float, dp_pct: float) -> float:
    def scale(hm, a_bt, a_thep, a_nc):
        return hm * (a_bt * k_bt + a_thep * k_thep + a_nc * k_nc)

    xl_new = (scale(th.get("hm1_KCN", 0), 0.42, 0.35, 0.23) +
              scale(th.get("hm2_MTM", 0), 0.50, 0.20, 0.30) +
              scale(th.get("hm3_MON", 0), 0.45, 0.25, 0.30) +
              scale(th.get("hm4_BMC", 0), 0.55, 0.15, 0.30) +
              th.get("hm5_PHU", 0))  # phụ trợ không đổi theo vật liệu
    overhead = dp_pct / 100 + 0.05 + 0.02 + 0.01 + 0.005
    return xl_new * (1 + overhead) * 1.10

# ══════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════

def _export_excel(pa_list: list, scores: dict) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    BLUE   = PatternFill("solid", fgColor="4472C4")
    YELLOW = PatternFill("solid", fgColor="FFE699")
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")

    def _hdr(ws, row, col, val, fill=None):
        c = ws.cell(row, col, val)
        c.font = Font(bold=True, color="FFFFFF" if fill == BLUE else "000000")
        if fill: c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        return c

    # ── Sheet 1: Tóm tắt ──
    ws = wb.active
    ws.title = "Tom tat"
    ws.merge_cells("A1:E1")
    ws["A1"] = "BẢNG SO SÁNH BA PHƯƠNG ÁN KẾT CẤU NHỊP"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = f"Ngày lập: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells("A2:E2")

    col_hdrs = ["Chỉ tiêu", "PA1 – Chi phí", "PA2 – Mỹ quan", "PA3 – AI KN", "Tốt nhất"]
    for ci, h in enumerate(col_hdrs, 1):
        _hdr(ws, 4, ci, h, BLUE)

    hm_keys  = ["hm1_KCN","hm2_MTM","hm3_MON","hm4_BMC","hm5_PHU",
                "tong_xay_lap","tong_dau_tu"]
    rows_data = [
        ("Loại dầm",             [p["loai_dam"]                     for p in pa_list], "text"),
        ("Công nghệ",            [p.get("cong_nghe","")             for p in pa_list], "text"),
        ("Chiều dài nhịp (m)",   [p["L_nhip"]                       for p in pa_list], "lower"),
        ("Số nhịp",              [p["n_nhip"]                       for p in pa_list], "lower"),
        ("Chiều cao dầm (m)",    [p["H_dam_m"]                      for p in pa_list], "neutral"),
        ("Bề rộng dầm (m)",      [p["B_dam_m"]                      for p in pa_list], "neutral"),
        ("Tỉ lệ L/H",            [p["LH_ratio"]                     for p in pa_list], "higher"),
        ("Số dầm/MCN",           [p["n_dam"]                        for p in pa_list], "lower"),
        ("Chi phí KCN (tỷ)",     [p["th"].get("hm1_KCN",0)/1e6     for p in pa_list], "lower"),
        ("Chi phí móng (tỷ)",    [p["th"].get("hm3_MON",0)/1e6     for p in pa_list], "lower"),
        ("TỔNG ĐẦU TƯ (tỷ)",    [p["tong_dau_tu"]/1e6              for p in pa_list], "lower"),
        ("Suất đầu tư (tr/m²)", [p["suat_dau_tu_trieu"]             for p in pa_list], "lower"),
        ("Điểm kỹ thuật (/60)",  [scores[p["key"]]["diem_ky_thuat"] for p in pa_list], "higher"),
        ("Điểm kinh tế (/30)",   [scores[p["key"]]["diem_kinh_te"]  for p in pa_list], "higher"),
        ("Điểm mỹ quan (/10)",   [scores[p["key"]]["diem_my_quan"]  for p in pa_list], "higher"),
        ("TỔNG ĐIỂM (/100)",     [scores[p["key"]]["tong_diem"]     for p in pa_list], "higher"),
    ]

    for ri, (label, vals, prefer) in enumerate(rows_data, 5):
        ws.cell(ri, 1, label).font = Font(bold=(ri >= 18))
        num_vals = [v for v in vals if isinstance(v, (int, float))]
        best_v = (min(num_vals) if prefer == "lower" else max(num_vals)) if num_vals else None
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(ri, ci)
            if isinstance(v, float):
                cell.value = round(v, 3)
                cell.number_format = "#,##0.000"
            else:
                cell.value = v
            if best_v is not None and v == best_v:
                cell.fill = GREEN
        if best_v is not None:
            best_pa = PA_KEYS[([v for _, vals_inner, _ in rows_data if vals_inner]
                                if False else vals).index(best_v)]
            ws.cell(ri, 5, PA_LABELS.get(best_pa, "")).font = Font(italic=True)

    ws.column_dimensions["A"].width = 28
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width = 20

    # ── Sheet 2: Chi tiết hạng mục ──
    ws2 = wb.create_sheet("Chi tiet hang muc")
    ws2["A1"] = "CHI TIẾT CHI PHÍ TỪNG HẠNG MỤC (tỷ VND)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.merge_cells("A1:E1")
    for ci, h in enumerate(["Hạng mục", "PA1", "PA2", "PA3", "Đơn vị"], 1):
        _hdr(ws2, 2, ci, h, BLUE)

    hm_detail = [
        ("I. Kết cấu nhịp (KCN)",         "hm1_KCN",      True),
        ("II. Mố trụ (MTM)",               "hm2_MTM",      False),
        ("III. Móng cọc",                  "hm3_MON",      False),
        ("IV. Bản mặt cầu + lớp phủ",     "hm4_BMC",      False),
        ("V. Phụ trợ (gối, khe, lan can)", "hm5_PHU",      False),
        ("▸ TỔNG XÂY LẮP",                "tong_xay_lap", True),
        ("  Dự phòng + TK + KS + TN + BH","du_phong",     False),
        ("  VAT 10%",                      "VAT",          False),
        ("▸ TỔNG ĐẦU TƯ (có VAT)",        "tong_dau_tu",  True),
    ]
    for ri, (label, key, bold) in enumerate(hm_detail, 3):
        c = ws2.cell(ri, 1, label)
        c.font = Font(bold=bold)
        for ci, pa in enumerate(pa_list, 2):
            val = pa["th"].get(key, 0) / 1e6
            cell = ws2.cell(ri, ci, round(val, 3))
            cell.number_format = "#,##0.000"
            if bold: cell.font = Font(bold=True)
        ws2.cell(ri, 5, "tỷ VND")

    for col in ["A","B","C","D","E"]:
        ws2.column_dimensions[col].width = 28

    # ── Sheet 3: Điểm đánh giá ──
    ws3 = wb.create_sheet("Diem danh gia")
    ws3["A1"] = "BẢNG ĐIỂM ĐÁNH GIÁ PHƯƠNG ÁN (100 điểm)"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3.merge_cells("A1:F1")
    for ci, h in enumerate(["Tiêu chí", "PA1", "PA2", "PA3", "Tối đa", "Ghi chú"], 1):
        _hdr(ws3, 2, ci, h, BLUE)

    criteria_rows = [
        ("──── NHÓM A: KỸ THUẬT (60đ) ────", None, None, True),
        ("A1 – L/H hợp lý",          "A1", 12, False),
        ("A2 – Nhịp phù hợp loại dầm","A2", 12, False),
        ("A3 – Tính khả thi thi công","A3", 12, False),
        ("A4 – Hiệu quả vật liệu",   "A4", 12, False),
        ("A5 – Độ bền lâu dài",      "A5", 12, False),
        ("──── NHÓM B: KINH TẾ (30đ) ────", None, None, True),
        ("B1 – Chi phí tổng",        "B1", 15, False),
        ("B2 – Suất đầu tư",         "B2", 10, False),
        ("B3 – Hiệu quả KT tổng hợp","B3", 5,  False),
        ("──── NHÓM C: MỸ QUAN (10đ) ────", None, None, True),
        ("C1 – Tỉ lệ thẩm mỹ L/H",  "C1", 5,  False),
        ("C2 – Kiểu dáng loại dầm",  "C2", 5,  False),
        ("══ TỔNG ĐIỂM ══",          "tong_diem", 100, True),
    ]
    for ri, (label, key, mx, is_hdr) in enumerate(criteria_rows, 3):
        c = ws3.cell(ri, 1, label)
        c.font = Font(bold=is_hdr)
        if is_hdr: c.fill = YELLOW
        if key:
            vals = [scores[pk].get(key, "") for pk in PA_KEYS]
            max_v = max((v for v in vals if isinstance(v, (int,float))), default=0)
            for ci, v in enumerate(vals, 2):
                cell = ws3.cell(ri, ci, v)
                if is_hdr: cell.font = Font(bold=True)
                if isinstance(v, (int,float)) and v == max_v:
                    cell.fill = GREEN
            ws3.cell(ri, 5, mx)

    for col in ["A","B","C","D","E","F"]:
        ws3.column_dimensions[col].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
# MAIN PAGE
# ══════════════════════════════════════════════════════════════════

st.title("📊 Dashboard so sánh ba phương án kết cấu nhịp")
st.caption(
    "Tổng hợp PA1 (tối ưu chi phí) · PA2 (tối ưu mỹ quan) · PA3 (AI khuyến nghị) "
    "từ Module 06 để hỗ trợ ra quyết định lựa chọn phương án cho hồ sơ TKCS."
)

# ── Load Module 12 ──
try:
    dt_mod = _load_du_toan_mod()
except Exception as e:
    st.error(f"Không thể tải Module 12 (DuToan_SoBo): {e}")
    st.stop()

# ── Determine data source ──
res      = st.session_state.get("res", {})
ss3pa    = st.session_state.get("so_sanh_3pa")

if ss3pa and isinstance(ss3pa, list) and len(ss3pa) == 3:
    pa_list = ss3pa
    _mode = "real"
else:
    _mode = "demo"
    st.info(
        "**Chế độ minh họa** — Chưa có dữ liệu 3 PA từ Module 06. "
        "Dashboard hiển thị: PA1 Super-T 38.2m · PA2 T ngược 18m×2 nhịp · PA3 Dầm I 33m."
    )
    with st.spinner("Đang tính dự toán ba phương án mẫu…"):
        pa_list = _demo_pa_list(dt_mod)
    if not pa_list or any(p["tong_dau_tu"] == 0 for p in pa_list):
        st.error("Lỗi tính toán dự toán mẫu. Kiểm tra lại Module 12.")
        st.stop()

# ── Score all PA ──
all_costs = [p["tong_dau_tu"] for p in pa_list]
scores    = {p["key"]: _score_pa(p, all_costs) for p in pa_list}
best_key  = max(scores, key=lambda k: scores[k]["tong_diem"])

# ════════════════════════════════════════════════════════════════
# ① THREE-COLUMN OVERVIEW
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("① Tóm tắt ba phương án")

cols = st.columns(3)
for pa, col in zip(pa_list, cols):
    pk  = pa["key"]
    clr = PA_COLORS[pk]
    sc  = scores[pk]
    th  = pa["th"]
    is_best = (pk == best_key)

    with col:
        badge_html = (
            '<br><span style="background:#f59e0b;color:white;padding:2px 8px;'
            'border-radius:12px;font-size:12px">🏆 Khuyến nghị</span>'
            if is_best else ""
        )
        border = f"3px solid #f59e0b" if is_best else f"1px solid {clr}44"
        st.markdown(
            f'<div style="background:{PA_BG[pk]};border:{border};border-radius:10px;'
            f'padding:12px;margin-bottom:6px">'
            f'<div style="color:{clr};font-weight:bold;font-size:15px">'
            f'{PA_LABELS[pk]}</div>'
            f'<span style="font-size:28px;font-weight:bold;color:{clr}">'
            f'{sc["tong_diem"]}</span>'
            f'<span style="color:#6b7280;font-size:13px"> / 100 điểm</span>'
            f'{badge_html}</div>',
            unsafe_allow_html=True,
        )

        # Cross-section drawing
        st.plotly_chart(_section_fig(pa, clr), use_container_width=True,
                        key=f"sec_{pk}", config={"displayModeBar": False})

        # Technical params
        st.markdown(f"**⚙️ Thông số kỹ thuật**")
        r1, r2 = st.columns(2)
        r1.metric("Loại dầm",    pa["loai_dam"])
        r2.metric("Công nghệ",   pa.get("cong_nghe","—")[:10])
        r1.metric("Chiều dài",   f"{pa['L_nhip']:.1f} m")
        r2.metric("Tổng nhịp",   pa["n_nhip"])
        r1.metric("Chiều cao",   f"{pa['H_dam_m']:.2f} m")
        r2.metric("L/H",         pa["LH_ratio"])
        r1.metric("Bề rộng dầm", f"{pa['B_dam_m']:.2f} m")
        r2.metric("Số dầm/MCN",  pa["n_dam"])

        # Cost
        st.markdown(f"**💰 Chi phí dự toán**")
        r1, r2 = st.columns(2)
        r1.metric("KCN",  f"{th.get('hm1_KCN',0)/1e6:.3f} tỷ")
        r2.metric("Móng", f"{th.get('hm3_MON',0)/1e6:.3f} tỷ")
        min_cost  = min(all_costs)
        diff_pct  = (pa["tong_dau_tu"] - min_cost) / max(min_cost, 1) * 100
        delta_str = (f"+{diff_pct:.1f}% so PA rẻ nhất" if diff_pct > 0.5 else "Rẻ nhất")
        st.metric("Tổng đầu tư",
                  f"{pa['tong_dau_tu']/1e6:.3f} tỷ",
                  delta_str, delta_color="inverse")
        st.metric("Suất đầu tư", f"{pa['suat_dau_tu_trieu']:.1f} tr/m²")

        # Score breakdown
        with st.expander(f"Điểm chi tiết — {sc['tong_diem']}/100"):
            score_rows = []
            for grp, sub_keys in [("A.Kỹ thuật",["A1","A2","A3","A4","A5"]),
                                   ("B.Kinh tế", ["B1","B2","B3"]),
                                   ("C.Mỹ quan", ["C1","C2"])]:
                for k in sub_keys:
                    score_rows.append({
                        "Nhóm": grp,
                        "Tiêu chí": SCORE_LABELS[k],
                        "Điểm": sc[k],
                    })
            st.dataframe(pd.DataFrame(score_rows),
                         use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════════════════════
# ② DETAILED COMPARISON TABLE
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("② Bảng so sánh chi tiết")

PA_SHORT = [p["key"] for p in pa_list]

def _best_worst_style(df: pd.DataFrame, prefer_map: dict):
    """Return CSS for each cell based on best/worst value."""
    colors = pd.DataFrame("", index=df.index, columns=df.columns)
    for _, row in df.iterrows():
        label = row.get("Tiêu chí", "")
        prefer = prefer_map.get(label, "neutral")
        if prefer == "neutral":
            continue
        num_vals = {}
        for pk in PA_SHORT:
            v = row.get(pk)
            if isinstance(v, (int, float)):
                num_vals[pk] = v
        if len(num_vals) < 2:
            continue
        best  = (min(num_vals.values()) if prefer == "lower"
                 else max(num_vals.values()))
        worst = (max(num_vals.values()) if prefer == "lower"
                 else min(num_vals.values()))
        idx = row.name
        for pk, v in num_vals.items():
            if v == best:
                colors.at[idx, pk] = "background-color:#dcfce7"
            elif v == worst:
                colors.at[idx, pk] = "background-color:#fee2e2"
    return colors

table_groups = [
    ("🏗️ Thông số kết cấu nhịp", [
        ("Loại dầm",           [p["loai_dam"] for p in pa_list],        "neutral"),
        ("Công nghệ DUL",      [p.get("cong_nghe","") for p in pa_list],"neutral"),
        ("Chiều dài nhịp (m)", [p["L_nhip"]  for p in pa_list],         "neutral"),
        ("Tổng số nhịp",       [p["n_nhip"]  for p in pa_list],         "lower"),
        ("Chiều cao dầm (m)",  [p["H_dam_m"] for p in pa_list],         "neutral"),
        ("Bề rộng dầm (m)",    [p["B_dam_m"] for p in pa_list],         "neutral"),
        ("Khoảng cách dầm (m)",[p["S_dam_m"] for p in pa_list],         "lower"),
        ("Số dầm / MCN",       [p["n_dam"]   for p in pa_list],         "lower"),
        ("Tỉ lệ L/H",          [p["LH_ratio"]for p in pa_list],         "higher"),
    ]),
    ("💰 Chi phí dự toán (tỷ VND)", [
        ("I. Kết cấu nhịp KCN",     [p["th"].get("hm1_KCN",0)/1e6  for p in pa_list], "lower"),
        ("II. Mố trụ MTM",           [p["th"].get("hm2_MTM",0)/1e6  for p in pa_list], "lower"),
        ("III. Móng cọc",            [p["th"].get("hm3_MON",0)/1e6  for p in pa_list], "lower"),
        ("IV. Bản mặt cầu + lớp phủ",[p["th"].get("hm4_BMC",0)/1e6 for p in pa_list], "lower"),
        ("V. Phụ trợ",               [p["th"].get("hm5_PHU",0)/1e6  for p in pa_list], "lower"),
        ("▸ Tổng xây lắp",           [p["th"].get("tong_xay_lap",0)/1e6 for p in pa_list], "lower"),
        ("▸ TỔNG ĐẦU TƯ",            [p["tong_dau_tu"]/1e6          for p in pa_list], "lower"),
        ("Suất đầu tư (tr/m²)",      [p["suat_dau_tu_trieu"]        for p in pa_list], "lower"),
    ]),
    ("📊 Đánh giá tổng hợp", [
        ("Điểm kỹ thuật (/60)", [scores[p["key"]]["diem_ky_thuat"] for p in pa_list], "higher"),
        ("Điểm kinh tế (/30)",  [scores[p["key"]]["diem_kinh_te"]  for p in pa_list], "higher"),
        ("Điểm mỹ quan (/10)",  [scores[p["key"]]["diem_my_quan"]  for p in pa_list], "higher"),
        ("TỔNG ĐIỂM (/100)",    [scores[p["key"]]["tong_diem"]     for p in pa_list], "higher"),
    ]),
]

for grp_title, rows in table_groups:
    st.markdown(f"##### {grp_title}")
    prefer_map = {label: pref for label, _, pref in rows}
    df_data = []
    for label, vals, _ in rows:
        row_dict = {"Tiêu chí": label}
        for pk, v in zip(PA_SHORT, vals):
            row_dict[pk] = round(v, 3) if isinstance(v, float) else v
        df_data.append(row_dict)
    df = pd.DataFrame(df_data)
    try:
        styled = df.style.apply(_best_worst_style, prefer_map=prefer_map, axis=None)
        st.dataframe(styled, use_container_width=True, hide_index=True)
    except Exception:
        st.dataframe(df, use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════════════════════
# ③ CHARTS
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("③ Biểu đồ so sánh trực quan")

ch_cols = st.columns(2)

# Chart 1: Radar
with ch_cols[0]:
    st.markdown("**Radar — Điểm theo nhóm tiêu chí**")
    radar_cats = ["Kỹ thuật\n(60)", "Kinh tế\n(30)", "Mỹ quan\n(10)",
                  "Thi công\n(max 55)", "Độ bền\n(max 55)"]
    fig_r = go.Figure()
    for pa in pa_list:
        pk = pa["key"]
        sc = scores[pk]
        vals = [sc["diem_ky_thuat"], sc["diem_kinh_te"], sc["diem_my_quan"],
                sc["A3"] * 4.5, sc["A5"] * 4.5]
        vals_c = vals + [vals[0]]
        cats_c = radar_cats + [radar_cats[0]]
        fig_r.add_trace(go.Scatterpolar(
            r=vals_c, theta=cats_c,
            fill="toself", fillcolor=PA_COLORS[pk] + "33",
            line=dict(color=PA_COLORS[pk], width=2),
            name=PA_LABELS[pk].split("—")[1].strip(),
        ))
    fig_r.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 62])),
        height=360, margin=dict(l=60, r=60, t=20, b=30),
        legend=dict(orientation="h", y=-0.15, font=dict(size=11)),
    )
    st.plotly_chart(fig_r, use_container_width=True, config={"displayModeBar": False})

# Chart 2: Grouped bar
with ch_cols[1]:
    st.markdown("**Cột nhóm — Chi phí từng hạng mục**")
    hm_names = ["KCN", "MTM", "MON", "BMC", "PHU"]
    hm_keys2 = ["hm1_KCN","hm2_MTM","hm3_MON","hm4_BMC","hm5_PHU"]
    fig_b = go.Figure()
    for pa in pa_list:
        fig_b.add_trace(go.Bar(
            name=pa["key"], x=hm_names,
            y=[pa["th"].get(k, 0) / 1e6 for k in hm_keys2],
            marker_color=PA_COLORS[pa["key"]],
        ))
    fig_b.update_layout(
        barmode="group", height=360,
        yaxis_title="tỷ VND",
        margin=dict(l=10, r=10, t=10, b=30),
        legend=dict(orientation="h", y=-0.15),
    )
    st.plotly_chart(fig_b, use_container_width=True, config={"displayModeBar": False})

ch_cols2 = st.columns(2)

# Chart 3: Stacked bar (% structure)
with ch_cols2[0]:
    st.markdown("**Cột xếp chồng — Cơ cấu chi phí (%)**")
    fig_s = go.Figure()
    for hi, (hm, key) in enumerate(zip(hm_names, hm_keys2)):
        pcts = []
        for pa in pa_list:
            xl  = pa["th"].get("tong_xay_lap", 1)
            pct = pa["th"].get(key, 0) / max(xl, 1) * 100
            pcts.append(round(pct, 1))
        fig_s.add_trace(go.Bar(
            name=hm,
            x=[PA_LABELS[pk].split("—")[1].strip() for pk in PA_KEYS],
            y=pcts,
        ))
    fig_s.update_layout(
        barmode="stack", height=360,
        yaxis_title="% / Tổng xây lắp",
        margin=dict(l=10, r=10, t=10, b=30),
        legend=dict(orientation="h", y=-0.2),
    )
    st.plotly_chart(fig_s, use_container_width=True, config={"displayModeBar": False})

# Chart 4: Scatter cost vs score
with ch_cols2[1]:
    st.markdown("**Scatter — Chi phí ↔ Điểm (phía trên-trái = tốt nhất)**")
    fig_sc = go.Figure()
    for pa in pa_list:
        pk = pa["key"]
        sc = scores[pk]
        fig_sc.add_trace(go.Scatter(
            x=[pa["tong_dau_tu"] / 1e6],
            y=[sc["tong_diem"]],
            mode="markers+text",
            text=[pa["key"]],
            textposition="top center",
            marker=dict(size=22, color=PA_COLORS[pk],
                        line=dict(width=2, color="white"),
                        symbol="circle"),
            name=PA_LABELS[pk],
        ))
    fig_sc.add_annotation(
        text="← Chi phí thấp, điểm cao = tối ưu",
        x=0.02, y=0.96, xref="paper", yref="paper",
        showarrow=False, font=dict(size=10, color="gray"),
    )
    fig_sc.update_layout(
        height=360,
        xaxis_title="Tổng đầu tư (tỷ VND) ← thấp hơn tốt",
        yaxis_title="Tổng điểm (/100) ↑ cao hơn tốt",
        margin=dict(l=10, r=10, t=10, b=30),
        showlegend=False,
    )
    st.plotly_chart(fig_sc, use_container_width=True, config={"displayModeBar": False})

# ════════════════════════════════════════════════════════════════
# ④ DECISION
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("④ Ra quyết định chọn phương án")

best_pa = next(p for p in pa_list if p["key"] == best_key)
best_sc = scores[best_key]

reasons = []
if best_sc["diem_ky_thuat"] == max(scores[pk]["diem_ky_thuat"] for pk in PA_KEYS):
    reasons.append("điểm kỹ thuật cao nhất")
if best_sc["diem_kinh_te"] == max(scores[pk]["diem_kinh_te"] for pk in PA_KEYS):
    reasons.append("chi phí tối ưu")
if best_sc["diem_my_quan"] == max(scores[pk]["diem_my_quan"] for pk in PA_KEYS):
    reasons.append("thẩm mỹ tốt nhất")
reason_str = " và ".join(reasons) if reasons else "tổng điểm cao nhất"

st.success(
    f"**Khuyến nghị tự động: {PA_LABELS[best_key]}**  \n"
    f"Lý do: {reason_str} (tổng điểm **{best_sc['tong_diem']}/100**).  \n"
    f"Thông số: {best_pa['loai_dam']} — L={best_pa['L_nhip']}m "
    f"— L/H={best_pa['LH_ratio']} — Suất ĐT {best_pa['suat_dau_tu_trieu']:.1f} tr/m²"
)

ghi_chu = st.text_area(
    "Ghi chú lý do chọn (nếu khác khuyến nghị):",
    key="_ghi_chu_pa",
    placeholder="Ví dụ: Chọn PA2 vì yêu cầu mỹ quan cao trong khu đô thị…",
    height=70,
)

dec_cols = st.columns(3)
for col, pa in zip(dec_cols, pa_list):
    pk = pa["key"]
    is_best = (pk == best_key)
    with col:
        if st.button(
            f"{'✅ ' if is_best else ''}Chọn {PA_LABELS[pk]}",
            key=f"choose_{pk}",
            type="primary" if is_best else "secondary",
            use_container_width=True,
        ):
            st.session_state["phuong_an_cuoi_cung"] = {
                "key":         pk,
                "ten_PA":      PA_LABELS[pk],
                "loai_dam":    pa["loai_dam"],
                "L_nhip":      pa["L_nhip"],
                "n_nhip":      pa["n_nhip"],
                "tong_dau_tu": pa["tong_dau_tu"],
                "suat":        pa["suat_dau_tu_trieu"],
                "tong_diem":   scores[pk]["tong_diem"],
                "timestamp":   datetime.now().isoformat(),
                "ghi_chu":     ghi_chu,
            }
            st.success(f"✅ Đã ghi nhận: {PA_LABELS[pk]}")
            st.rerun()

pac = st.session_state.get("phuong_an_cuoi_cung")
if pac:
    ts = pac["timestamp"][:16].replace("T", " ")
    st.info(
        f"**Phương án đã chọn:** {pac['ten_PA']} — {pac['loai_dam']} L={pac['L_nhip']}m "
        f"| Tổng điểm: {pac['tong_diem']}/100 | Suất ĐT: {pac['suat']:.1f} tr/m² "
        f"| Chọn lúc: {ts}"
    )

# ════════════════════════════════════════════════════════════════
# ⑤ SENSITIVITY ANALYSIS
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("⑤ Phân tích nhạy cảm")

with st.expander("🎚️ Điều chỉnh hệ số giá và xem tác động lên tổng chi phí"):
    sl_cols = st.columns(4)
    k_bt   = sl_cols[0].slider("Hệ số BT", 0.80, 1.30, 1.00, 0.05, key="k_bt",
                                help="Điều chỉnh đơn giá bê tông")
    k_thep = sl_cols[1].slider("Hệ số thép", 0.80, 1.30, 1.00, 0.05, key="k_thep",
                                help="Điều chỉnh đơn giá cốt thép")
    k_nc   = sl_cols[2].slider("Hệ số NC+Máy", 0.80, 1.30, 1.00, 0.05, key="k_nc",
                                help="Điều chỉnh nhân công và máy thi công")
    dp_pct = sl_cols[3].slider("Dự phòng (%)", 5, 20, 10, 1, key="dp_sa",
                                help="Hệ số dự phòng rủi ro")

    sc_cols = st.columns(3)
    if sc_cols[0].button("🟢 Lạc quan (−15%)", key="sa_opt"):
        for k, v in [("k_bt",0.85),("k_thep",0.85),("k_nc",0.85),("dp_sa",7)]:
            st.session_state[k] = v
        st.rerun()
    if sc_cols[1].button("🟡 Trung tính (gốc)", key="sa_base"):
        for k, v in [("k_bt",1.00),("k_thep",1.00),("k_nc",1.00),("dp_sa",10)]:
            st.session_state[k] = v
        st.rerun()
    if sc_cols[2].button("🔴 Bi quan (+20%)", key="sa_pess"):
        for k, v in [("k_bt",1.15),("k_thep",1.20),("k_nc",1.10),("dp_sa",15)]:
            st.session_state[k] = v
        st.rerun()

    # Recompute costs
    k_bt   = st.session_state.get("k_bt",   1.00)
    k_thep = st.session_state.get("k_thep", 1.00)
    k_nc   = st.session_state.get("k_nc",   1.00)
    dp_pct = st.session_state.get("dp_sa",  10)

    sa_cols = st.columns(3)
    for col, pa in zip(sa_cols, pa_list):
        pk        = pa["key"]
        base_cost = pa["tong_dau_tu"]
        new_cost  = _sensitivity_total(pa["th"], k_bt, k_thep, k_nc, dp_pct)
        delta_pct = (new_cost - base_cost) / max(base_cost, 1) * 100
        with col:
            st.metric(
                f"{PA_LABELS[pk].split('—')[1].strip()} (điều chỉnh)",
                f"{new_cost/1e6:.3f} tỷ",
                f"{delta_pct:+.1f}% so gốc",
                delta_color="inverse",
            )

    fig_sa = go.Figure()
    fig_sa.add_trace(go.Bar(
        name="Gốc (tham chiếu)",
        x=[PA_LABELS[pk] for pk in PA_KEYS],
        y=[p["tong_dau_tu"] / 1e6 for p in pa_list],
        marker_color=["#93c5fd", "#c4b5fd", "#86efac"],
        opacity=0.7,
    ))
    fig_sa.add_trace(go.Bar(
        name="Sau điều chỉnh",
        x=[PA_LABELS[pk] for pk in PA_KEYS],
        y=[_sensitivity_total(p["th"], k_bt, k_thep, k_nc, dp_pct) / 1e6
           for p in pa_list],
        marker_color=list(PA_COLORS.values()),
    ))
    fig_sa.update_layout(
        barmode="group", height=300,
        yaxis_title="tỷ VND",
        margin=dict(l=10, r=10, t=10, b=10),
        legend=dict(orientation="h", y=-0.35),
    )
    st.plotly_chart(fig_sa, use_container_width=True, config={"displayModeBar": False})

# ════════════════════════════════════════════════════════════════
# ⑥ EXPORT
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("⑥ Xuất hồ sơ so sánh")

exp_cols = st.columns(3)

with exp_cols[0]:
    excel_bytes = _export_excel(pa_list, scores)
    if excel_bytes:
        st.download_button(
            "📥 Xuất Excel (3 sheets)",
            data=excel_bytes,
            file_name=f"SoSanh_3PA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.warning("Cần cài `openpyxl` để xuất Excel: `pip install openpyxl`")

with exp_cols[1]:
    # Markdown report
    md_parts = [
        "# Báo cáo So sánh Ba Phương án Kết cấu Nhịp\n",
        f"**Ngày lập:** {datetime.now().strftime('%d/%m/%Y %H:%M')}\n",
        "## Tóm tắt\n",
        "| Chỉ tiêu | PA1 | PA2 | PA3 |",
        "| :--- | :---: | :---: | :---: |",
    ]
    for label, vals in [
        ("Loại dầm",          [p["loai_dam"]                      for p in pa_list]),
        ("Công nghệ",         [p.get("cong_nghe","")              for p in pa_list]),
        ("Chiều dài nhịp",   [f"{p['L_nhip']:.1f} m"             for p in pa_list]),
        ("Tổng số nhịp",     [str(p["n_nhip"])                   for p in pa_list]),
        ("Tỉ lệ L/H",        [str(p["LH_ratio"])                  for p in pa_list]),
        ("Tổng đầu tư",      [f"{p['tong_dau_tu']/1e6:.3f} tỷ"  for p in pa_list]),
        ("Suất đầu tư",      [f"{p['suat_dau_tu_trieu']:.1f} tr/m²" for p in pa_list]),
        ("Tổng điểm (/100)", [str(scores[p["key"]]["tong_diem"]) for p in pa_list]),
    ]:
        md_parts.append(f"| {label} | {' | '.join(str(v) for v in vals)} |")
    md_parts += [
        f"\n## Khuyến nghị\n",
        f"**{PA_LABELS[best_key]}** — Tổng điểm {best_sc['tong_diem']}/100\n",
        f"Lý do: {reason_str}\n",
    ]
    md_report = "\n".join(md_parts)
    st.download_button(
        "📄 Báo cáo Markdown",
        data=md_report.encode("utf-8"),
        file_name=f"BaoCao_SoSanh_{datetime.now().strftime('%Y%m%d_%H%M')}.md",
        mime="text/markdown",
        use_container_width=True,
    )

with exp_cols[2]:
    dump = {
        "timestamp": datetime.now().isoformat(),
        "khuyen_nghi": best_key,
        "phuong_an": [
            {k: v for k, v in p.items() if k not in ("beam_params","th","pa_dt")}
            for p in pa_list
        ],
        "diem_danh_gia": {pk: {k: int(v) if isinstance(v, (int,float)) else v
                               for k,v in sc.items()}
                          for pk, sc in scores.items()},
    }
    st.download_button(
        "📊 Dữ liệu JSON",
        data=json.dumps(dump, ensure_ascii=False, indent=2).encode("utf-8"),
        file_name=f"SS_data_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
        mime="application/json",
        use_container_width=True,
    )

# ════════════════════════════════════════════════════════════════
# ⑦ HISTORY
# ════════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("⑦ Lịch sử phiên so sánh")

if "lich_su_so_sanh" not in st.session_state:
    st.session_state["lich_su_so_sanh"] = []

if st.button("💾 Lưu phiên so sánh hiện tại", key="save_history"):
    entry = {
        "timestamp":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "khuyen_nghi": best_key,
        "pa_chon":     st.session_state.get("phuong_an_cuoi_cung", {}).get("key", ""),
        "ghi_chu":     st.session_state.get("_ghi_chu_pa", ""),
        "pa_summary": [
            {
                "key":        p["key"],
                "loai_dam":   p["loai_dam"],
                "L_nhip":     p["L_nhip"],
                "tong_dau_tu":p["tong_dau_tu"],
                "tong_diem":  scores[p["key"]]["tong_diem"],
            }
            for p in pa_list
        ],
    }
    st.session_state["lich_su_so_sanh"].append(entry)
    st.success("Đã lưu phiên so sánh!")

history = st.session_state.get("lich_su_so_sanh", [])
if history:
    hist_rows = []
    for i, e in enumerate(reversed(history), 1):
        row = {
            "#":           len(history) - i + 1,
            "Thời gian":   e["timestamp"],
            "Khuyến nghị": PA_LABELS.get(e.get("khuyen_nghi",""), e.get("khuyen_nghi","")),
            "Đã chọn":     PA_LABELS.get(e.get("pa_chon",""), e.get("pa_chon","")) or "—",
            "Ghi chú":     e.get("ghi_chu",""),
        }
        for ps in e.get("pa_summary",[]):
            row[f"{ps['key']} ({ps['loai_dam']})"] = (
                f"{ps['tong_dau_tu']/1e6:.2f} tỷ | {ps['tong_diem']}đ"
            )
        hist_rows.append(row)
    st.dataframe(pd.DataFrame(hist_rows), use_container_width=True, hide_index=True)
else:
    st.info("Chưa có lịch sử. Nhấn **Lưu phiên so sánh** để bắt đầu theo dõi.")
